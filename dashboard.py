# -*- coding: utf-8 -*-
"""
학교급식 시장규모 대시보드 서버
실행: py -X utf8 dashboard.py
접속: http://localhost:8765
"""
import sys, os, subprocess, json, glob, re, threading
from datetime import datetime

IS_CLOUD = bool(os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RAILWAY_PROJECT_ID"))

def ensure_flask():
    try:
        import flask
    except ImportError:
        print("[설치 중] flask...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "flask", "-q"])

ensure_flask()

from flask import Flask, jsonify, send_file, request, after_this_request

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
# 클라우드(Railway)에서는 영구 볼륨(/app/data)에 log/output을 두어
# 재배포 후에도 재수집 결과가 유지되도록 함. 로컬은 기존 경로 그대로.
DATA_DIR     = "/app/data" if IS_CLOUD else SCRIPT_DIR
LOG_DIR      = os.path.join(DATA_DIR, "log")
OUTPUT_DIR   = os.path.join(DATA_DIR, "output")
SKILL_PY     = os.path.join(SCRIPT_DIR, "학교급식규모.py")

# ── 분석 기준 연도 (매년 1월에 수동 업데이트) ────────────────
ANALYSIS_YEAR = 2025   # 항상 이 연도 vs 전년도(2024) 비교

VACATION    = {1, 7, 12}
MONTH_NAMES = ["1월","2월","3월","4월","5월","6월",
               "7월","8월","9월","10월","11월","12월"]

app = Flask(__name__)
running_tasks = {}   # keyword → "running" | "done" | "error:..."
bluesis_tasks = {}   # keyword → "running" | "done" | "error:..."  (로컬 전용)

COLLECTOR_PY = os.path.join(SCRIPT_DIR, "bluesis_collector.py")


def _bluesis_needs_update(product_info: list) -> bool:
    if not product_info:
        return False
    return any(not p.get("bluesis_ingredients") for p in product_info)


def _start_bluesis_bg(keyword: str, year: int):
    """로컬 전용: subprocess로 bluesis_collector.py를 실행해 백그라운드 수집."""
    if IS_CLOUD:          # Railway에서는 실행 안 함
        return
    if bluesis_tasks.get(keyword) == "running":
        return
    if not os.path.exists(COLLECTOR_PY):
        return

    def _run():
        bluesis_tasks[keyword] = "running"
        try:
            ret = subprocess.run(
                [sys.executable, "-X", "utf8", COLLECTOR_PY,
                 keyword, str(year), LOG_DIR],
                cwd=SCRIPT_DIR, timeout=300
            )
            bluesis_tasks[keyword] = "done" if ret.returncode == 0 else f"error:{ret.returncode}"
            print(f"[bluesis_bg] {keyword} {'완료' if ret.returncode == 0 else '오류'}")
        except Exception as e:
            bluesis_tasks[keyword] = f"error:{e}"
            print(f"[bluesis_bg] {keyword} 예외: {e}")

    threading.Thread(target=_run, daemon=True).start()
    print(f"[bluesis_bg] {keyword} 백그라운드 수집 시작 (로컬)")


# ── 유틸 ──────────────────────────────────────────────────
def get_latest_log(keyword: str, year: int):
    pattern = os.path.join(LOG_DIR, f"{keyword}_학교급식_{year}_*.json")
    files   = sorted(glob.glob(pattern))
    return files[-1] if files else None


def read_product_from_excel(keyword: str, year: int) -> list:
    """Excel 경쟁사_제품정보 Sheet에서 제품 정보를 읽어 반환.
    레이블(B열)을 동적으로 읽어 구버전/신버전 Sheet 모두 지원."""
    try:
        import openpyxl, re as _re
    except ImportError:
        return []

    def _find_excel(yr):
        pat = os.path.join(OUTPUT_DIR, f"{keyword}_학교급식규모_{yr}_*.xlsx")
        return [f for f in sorted(glob.glob(pat)) if not os.path.basename(f).startswith("~$")]

    files = _find_excel(year) or _find_excel(year - 1)
    if not files:
        return []

    try:
        wb = openpyxl.load_workbook(files[-1], data_only=True)
        if "경쟁사_제품정보" not in wb.sheetnames:
            return []
        ws = wb["경쟁사_제품정보"]

        # B열 레이블 → 행번호 매핑 (구버전/신버전 모두 지원)
        row_map = {}
        for row in range(1, 15):
            lbl = str(ws.cell(row=row, column=2).value or "").strip()
            if "업체명"   in lbl: row_map["company"]     = row
            if "제품명"   in lbl: row_map["product"]     = row
            if "규격"    in lbl or "중량" in lbl: row_map["standard"] = row
            if "방학제외" in lbl: row_map["mavg"]        = row
            if "원재료"   in lbl: row_map["ingredients"] = row
            if "원산지"   in lbl: row_map["origin"]      = row  # 구버전
            if "블루시스" in lbl and "단가" in lbl: row_map["kprice"] = row

        r_co   = row_map.get("company",     4)
        r_prod = row_map.get("product",     5)
        r_std  = row_map.get("standard",    6)
        r_mavg = row_map.get("mavg",        7)
        r_ingr = row_map.get("ingredients", 8)
        r_orig = row_map.get("origin",      None)
        r_kp   = row_map.get("kprice",     10)

        result = []
        for col in range(3, 9):
            company = str(ws.cell(row=r_co, column=col).value or "").strip()
            if not company or company in ("항목", "FoodnBid"):
                break

            product     = str(ws.cell(row=r_prod, column=col).value or "").strip()
            standard    = str(ws.cell(row=r_std,  column=col).value or "확인 필요").strip()
            mavg_raw    = str(ws.cell(row=r_mavg, column=col).value or "0").strip()
            ingredients = str(ws.cell(row=r_ingr, column=col).value or "").strip()
            origin      = str(ws.cell(row=r_orig, column=col).value or "").strip() if r_orig else ""
            bluesis     = str(ws.cell(row=r_kp,   column=col).value or "직접 확인").strip()

            m = _re.search(r"([\d,]+)", mavg_raw)
            monthly_avg = int(m.group(1).replace(",", "")) if m else 0

            # 구버전 placeholder 텍스트는 빈값으로 처리
            if any(kw in ingredients for kw in ("직접 기재", "직접 확인", "주원료(")):
                ingredients = ""

            result.append({
                "rank":                  col - 2,
                "company":               company,
                "product":               product,
                "annual":                0,
                "monthly_avg":           monthly_avg,
                "bluesis_standard":      standard,
                "bluesis_ingredients":   ingredients,
                "bluesis_kprice":        bluesis,
                "origin":                origin,
            })
        wb.close()
        return result
    except Exception as e:
        print(f"[경고] Excel 제품정보 읽기 실패: {e}")
        return []


def find_best_year(keyword: str) -> int:
    """ANALYSIS_YEAR 기준 연도 로그가 있으면 반환, 없으면 ANALYSIS_YEAR 고정."""
    for yr in [ANALYSIS_YEAR, ANALYSIS_YEAR - 1]:
        if get_latest_log(keyword, yr):
            return ANALYSIS_YEAR if get_latest_log(keyword, ANALYSIS_YEAR) else yr
    return ANALYSIS_YEAR


def compute_data(keyword: str):
    """키워드만으로 현재연도+전년도 데이터를 자동 결합해 반환."""
    year = find_best_year(keyword)
    curr_path = get_latest_log(keyword, year)
    if not curr_path:
        return None

    with open(curr_path, encoding="utf-8") as f:
        curr = json.load(f)

    prev_path = get_latest_log(keyword, year - 1)
    prev_raw  = None
    if prev_path:
        with open(prev_path, encoding="utf-8") as f:
            prev_raw = json.load(f)

    results     = curr["results"]
    annual      = curr["annual_total"]
    prev_annual = prev_raw["annual_total"] if prev_raw else 0

    # 시장규모 환산
    COV, SU, DAN = 0.60, 0.405, 0.225
    def market(a):
        if not a:
            return {"foodnbid": 0, "school_suwon": 0, "school_nation": 0, "total": 0}
        sw = int(a / COV)
        sn = int(sw / SU)
        return {"foodnbid": a, "school_suwon": sw, "school_nation": sn, "total": int(sn * (1 + DAN))}

    curr_mkt = market(annual)
    prev_mkt = market(prev_annual)

    # 전년 월별 맵
    prev_map = {}
    if prev_raw:
        for r in prev_raw["results"]:
            prev_map[r["month"]] = r["total"]

    # 방학월(1·7·12월) 제외 합계 — 방학월은 매출 변동폭이 비정상적으로 커서
    # 전년대비 증감률(%) 계산에서만 제외 (연간 실측 총액 자체는 방학월 포함)
    annual_ex      = sum(r["total"] for r in results if r["month"] not in VACATION)
    prev_annual_ex = sum(prev_map.get(m, 0) for m in range(1, 13) if m not in VACATION)
    yoy_ex = round((annual_ex - prev_annual_ex) / prev_annual_ex * 100, 1) if prev_annual_ex else None

    # 차년도(진행중) 월별 맵 — 있으면 로드 (예: 2026년 1~5월)
    next_path = get_latest_log(keyword, year + 1)
    next_raw  = None
    if next_path:
        with open(next_path, encoding="utf-8") as f:
            next_raw = json.load(f)

    next_map = {}
    if next_raw:
        for r in next_raw["results"]:
            next_map[r["month"]] = r["total"]

        # 진행 중인 이번 달은 낙찰 데이터가 미완결 상태이므로 집계에서 제외
        if year + 1 == datetime.now().year:
            cur_month = datetime.now().month
            next_map = {m: v for m, v in next_map.items() if m < cur_month}

    # 차년도 연간 추정 — 계절성 비중(과거 동기간 평균 비중) 반영 연환산
    y2026 = None
    if next_map:
        ytd_months = max((m for m, v in next_map.items() if v > 0), default=0)
        if ytd_months:
            sum_next     = sum(next_map.get(m, 0) for m in range(1, ytd_months + 1))
            sum_curr_ytd = sum(r["total"] for r in results if r["month"] <= ytd_months)
            sum_prev_ytd = sum(prev_map.get(m, 0) for m in range(1, ytd_months + 1))

            # 1~ytd_months월이 연간에서 차지하는 비중 (당해/전년 평균)
            w_curr  = sum_curr_ytd / annual if annual else 0
            w_prev  = sum_prev_ytd / prev_annual if prev_annual else 0
            weights = [w for w in (w_curr, w_prev) if w > 0]
            w_avg   = sum(weights) / len(weights) if weights else 0

            est_annual = int(sum_next / w_avg) if w_avg else 0

            # 방학월(1·7·12월) 제외 — 등락폭(%) 계산은 방학월을 빼고 산정
            sum_next_ex     = sum(next_map.get(m, 0) for m in range(1, ytd_months + 1) if m not in VACATION)
            sum_curr_ytd_ex = sum(r["total"] for r in results if r["month"] <= ytd_months and r["month"] not in VACATION)
            sum_prev_ytd_ex = sum(prev_map.get(m, 0) for m in range(1, ytd_months + 1) if m not in VACATION)

            w_curr_ex  = sum_curr_ytd_ex / annual_ex if annual_ex else 0
            w_prev_ex  = sum_prev_ytd_ex / prev_annual_ex if prev_annual_ex else 0
            weights_ex = [w for w in (w_curr_ex, w_prev_ex) if w > 0]
            w_avg_ex   = sum(weights_ex) / len(weights_ex) if weights_ex else 0

            est_annual_ex = int(sum_next_ex / w_avg_ex) if w_avg_ex else 0
            ytd_yoy = round((sum_next_ex - sum_curr_ytd_ex) / sum_curr_ytd_ex * 100, 1) if sum_curr_ytd_ex else None
            est_yoy = round((est_annual_ex - annual_ex) / annual_ex * 100, 1) if annual_ex and est_annual_ex else None

            y2026 = {
                "year":         year + 1,
                "ytd_months":   ytd_months,
                "sum_ytd":      sum_next,
                "sum_curr_ytd": sum_curr_ytd,
                "ytd_yoy":      ytd_yoy,
                "weight_pct":   round(w_avg_ex * 100, 1),
                "weight_curr":  round(w_curr_ex * 100, 1),
                "weight_prev":  round(w_prev_ex * 100, 1),
                "est_annual":   est_annual,
                "est_yoy":      est_yoy,
                # 산출식 상세 표시용 (방학월 1·7·12월 제외 기준 원자료)
                "sum_next_ex":     sum_next_ex,
                "sum_curr_ytd_ex": sum_curr_ytd_ex,
                "sum_prev_ytd_ex": sum_prev_ytd_ex,
                "est_annual_ex":   est_annual_ex,
            }

    # 월별 데이터
    best_month = max(results, key=lambda x: x["total"])["month"]
    monthly = []
    for r in results:
        prev_m = prev_map.get(r["month"], 0)
        yoy    = round((r["total"] - prev_m) / prev_m * 100, 1) if prev_m else None
        next_m = next_map.get(r["month"], 0)
        yoy_next = round((next_m - r["total"]) / r["total"] * 100, 1) if next_m and r["total"] else None
        monthly.append({
            "month":       r["month"],
            "label":       MONTH_NAMES[r["month"] - 1],
            "current":     r["total"],
            "prev":        prev_m,
            "yoy":         yoy,
            "next":        next_m or None,
            "yoy_next":    yoy_next,
            "is_vacation": r["month"] in VACATION,
            "is_best":     r["month"] == best_month,
            "top3":        r.get("companies", [])[:3],
        })

    # 경쟁사 데이터
    comp_annual  = {}
    comp_monthly = {}
    for r in results:
        for co in r.get("companies", []):
            nm, amt = co["company"], co["amount"]
            comp_annual[nm]  = comp_annual.get(nm, 0) + amt
            comp_monthly.setdefault(nm, {})[r["month"]] = amt

    sorted_comps = sorted(comp_annual.items(), key=lambda x: -x[1])

    def excl_vac_avg(company):
        months = [comp_monthly.get(company, {}).get(m, 0)
                  for m in range(1, 13)
                  if m not in VACATION and comp_monthly.get(company, {}).get(m, 0) > 0]
        return int(sum(months) / len(months)) if months else 0

    comp_list = [{
        "company":              c,
        "annual":               a,
        "share":                round(a / annual * 100, 1) if annual else 0,
        "monthly_avg_excl_vac": excl_vac_avg(c),
        "monthly":              {str(m): comp_monthly.get(c, {}).get(m, 0) for m in range(1, 13)},
    } for c, a in sorted_comps[:15]]

    # 제품 정보: JSON 우선, 없으면 Excel Sheet4 fallback
    product_info = []
    for pf in sorted(glob.glob(os.path.join(LOG_DIR, f"{keyword}_제품정보_{year}_*.json"))):
        with open(pf, encoding="utf-8") as f:
            product_info = json.load(f)

    if not product_info:
        product_info = read_product_from_excel(keyword, year)

    # annual/bluesis 필드 보완
    if product_info:
        comp_by_name   = {c["company"]: c for c in comp_list}
        excel_info_map = {p["company"]: p for p in read_product_from_excel(keyword, year)}
        for p in product_info:
            co = comp_by_name.get(p["company"], {})
            if co:
                if not p.get("annual"):
                    p["annual"] = co["annual"]
                if not p.get("monthly_avg"):
                    p["monthly_avg"] = co["monthly_avg_excl_vac"]
            ep = excel_info_map.get(p["company"], {})
            # 구버전 JSON → Excel에서 새 필드 보완
            for fld in ("bluesis_kprice", "bluesis_standard", "bluesis_ingredients",
                        "bluesis_image_b64"):
                if not p.get(fld):
                    p[fld] = ep.get(fld, "")
            # 하위 호환: 예전 bluesis_price 필드
            if not p.get("bluesis_price"):
                p["bluesis_price"] = p.get("bluesis_kprice", ep.get("bluesis_price", ""))

    return {
        "keyword":       keyword,
        "year":          year,
        "prev_year":     year - 1,
        "next_year":     year + 1,
        "annual":        annual,
        "prev_annual":   prev_annual,
        "annual_ex":     annual_ex,
        "prev_annual_ex": prev_annual_ex,
        "yoy_ex":        yoy_ex,
        "best_month":    best_month,
        "collected_at":  curr.get("collected_at", ""),
        "market_curr":   curr_mkt,
        "market_prev":   prev_mkt,
        "monthly":       monthly,
        "competitors":   comp_list,
        "product_info":  product_info,
        "has_prev":      prev_raw is not None,
        "y2026":         y2026,
        "exclude_keyword": curr.get("methodology", {}).get("exclude_keyword", ""),
    }


# ── API ───────────────────────────────────────────────────
@app.route("/")
def index():
    return send_file(os.path.join(SCRIPT_DIR, "dashboard.html"))


@app.route("/api/search")
def api_search():
    keyword = request.args.get("keyword", "").strip()
    if not keyword:
        return jsonify({"error": "keyword_required"}), 400
    data = compute_data(keyword)
    if data is None:
        return jsonify({"error": "not_found",
                        "message": f'"{keyword}" 데이터가 없습니다. 수집을 시작해 주세요.'}), 404

    # 로컬 전용: 블루시스 상세 없으면 자동 백그라운드 수집 시작
    if not IS_CLOUD and _bluesis_needs_update(data.get("product_info", [])):
        _start_bluesis_bg(keyword, data["year"])
        data["bluesis_collecting"] = True
    else:
        data["bluesis_collecting"] = (not IS_CLOUD and bluesis_tasks.get(keyword) == "running")

    return jsonify(data)


@app.route("/api/bluesis_status")
def api_bluesis_status():
    """로컬 전용: 블루시스 백그라운드 수집 상태 + 완료 시 최신 데이터 반환"""
    keyword = request.args.get("keyword", "").strip()
    if IS_CLOUD:
        return jsonify({"status": "unavailable"})
    status = bluesis_tasks.get(keyword, "idle")
    if status == "done":
        data = compute_data(keyword)
        return jsonify({"status": "done", "data": data})
    return jsonify({"status": status})


@app.route("/api/cache")
def api_cache():
    """캐시된 키워드 목록 — ANALYSIS_YEAR(2025) 또는 전년도(2024) 파일만 표시"""
    valid_years = {ANALYSIS_YEAR, ANALYSIS_YEAR - 1}
    files = glob.glob(os.path.join(LOG_DIR, "*_학교급식_*_*.json"))
    seen, result = set(), []
    # ANALYSIS_YEAR 파일 우선 (역순 정렬 → 2025 > 2024 순서)
    for f in sorted(files, reverse=True):
        m = re.match(r"(.+)_학교급식_(\d{4})_", os.path.basename(f))
        if m:
            kw, yr = m.group(1), int(m.group(2))
            if yr not in valid_years:
                continue
            if kw not in seen:
                seen.add(kw)
                # 항상 ANALYSIS_YEAR 기준으로 표시
                result.append({"keyword": kw, "year": ANALYSIS_YEAR})
    return jsonify(result)


@app.route("/api/run")
def api_run():
    keyword = request.args.get("keyword", "").strip()
    exclude = request.args.get("exclude", "").strip()
    if not keyword:
        return jsonify({"error": "keyword_required"}), 400

    if running_tasks.get(keyword) == "running":
        return jsonify({"status": "already_running"})

    # 수집 연도: 항상 ANALYSIS_YEAR(2025) 기준 — 스킬이 2025+2024 자동 수집
    year_str = str(ANALYSIS_YEAR)

    def _run():
        running_tasks[keyword] = "running"
        try:
            subprocess.run(
                [sys.executable, "-X", "utf8", SKILL_PY, keyword, year_str, exclude],
                cwd=SCRIPT_DIR
            )
            running_tasks[keyword] = "done"
        except Exception as e:
            running_tasks[keyword] = f"error:{e}"

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "started", "keyword": keyword, "year": year_str, "exclude": exclude})


def _add_y2026_sheet(wb, data):
    """{next_year}년(진행중) 데이터를 별도 시트로 추가/갱신합니다.
    Sheet1의 차트·컬럼 구조는 건드리지 않고 새 시트만 추가합니다."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    y2026     = data["y2026"]
    keyword   = data["keyword"]
    year      = data["year"]
    next_year = data["next_year"]
    monthly   = data["monthly"]
    annual_ex = data["annual_ex"]

    def fmt(n):
        return f"{n:,}"

    def fmt_eok(n):
        return f"{n / 1e8:.2f}억원"

    def yoy_str(v):
        if v is None:
            return "-"
        arrow = "▲" if v >= 0 else "▼"
        return f"{arrow}{abs(v):.1f}%"

    sheet_name = f"{next_year}년_진행현황"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    def side(): return Side(style="thin", color="CCCCCC")
    BD = Border(left=side(), right=side(), top=side(), bottom=side())
    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    P_BLUE  = PatternFill("solid", fgColor="1F4E79")
    P_GREEN = PatternFill("solid", fgColor="E8F5E9")
    P_VAC   = PatternFill("solid", fgColor="FFF3E0")
    P_GRAY  = PatternFill("solid", fgColor="F0F0F0")

    FT   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
    FH   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    FB   = Font(name="맑은 고딕", bold=True, size=10)
    FN   = Font(name="맑은 고딕", size=10)
    FSRC = Font(name="맑은 고딕", size=8, italic=True, color="888888")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for col in "CDEF":
        ws.column_dimensions[col].width = 17

    r = 1
    ws.merge_cells(f"B{r}:F{r}")
    c = ws.cell(r, 2, f"【{keyword}】 {next_year}년 진행현황  (vs {year}년 동월 비교)")
    c.fill = P_BLUE; c.font = FT; c.alignment = AC; c.border = BD
    ws.row_dimensions[r].height = 32
    r += 1

    headers = ["월", f"{year}년 실측(원)", f"{next_year}년 실측(원)", "동월대비", "비고"]
    for j, h in enumerate(headers, 2):
        cell = ws.cell(r, j, h)
        cell.fill = P_BLUE; cell.font = FH; cell.alignment = AC; cell.border = BD
    ws.row_dimensions[r].height = 24
    r += 1

    for m in monthly:
        is_vac = m["is_vacation"]
        ws.cell(r, 2, m["label"])
        ws.cell(r, 3, m["current"]).number_format = "#,##0"
        if m["next"] is not None:
            ws.cell(r, 4, m["next"]).number_format = "#,##0"
        else:
            ws.cell(r, 4, "-")
        if m["yoy_next"] is not None:
            yc = ws.cell(r, 5, m["yoy_next"] / 100)
            yc.number_format = "+0.0%;-0.0%"
        else:
            ws.cell(r, 5, "-")
        ws.cell(r, 6, "방학월 (등락률 산정 제외)" if is_vac else "")
        for col in range(2, 7):
            cell = ws.cell(r, col)
            cell.border = BD
            cell.font = FN
            cell.alignment = AC if col != 6 else AL
            if is_vac:
                cell.fill = P_VAC
        r += 1

    # YTD 합계 (방학월 제외)
    ws.cell(r, 2, f"1~{y2026['ytd_months']}월 합계\n(방학월 제외)")
    ws.cell(r, 3, y2026["sum_curr_ytd_ex"]).number_format = "#,##0"
    ws.cell(r, 4, y2026["sum_next_ex"]).number_format = "#,##0"
    if y2026["ytd_yoy"] is not None:
        yc = ws.cell(r, 5, y2026["ytd_yoy"] / 100)
        yc.number_format = "+0.0%;-0.0%"
    else:
        ws.cell(r, 5, "-")
    ws.cell(r, 6, "YTD 실성장률")
    for col in range(2, 7):
        cell = ws.cell(r, col)
        cell.fill = P_GREEN; cell.font = FB; cell.border = BD
        cell.alignment = AC if col != 6 else AL
    r += 2

    # ── 연간 추정 산출 내역
    ws.merge_cells(f"B{r}:F{r}")
    c = ws.cell(r, 2, f"{next_year}년 연간 추정 산출 내역  (방학월 1·7·12월 제외 기준)")
    c.fill = P_BLUE; c.font = FH; c.alignment = AC; c.border = BD
    r += 1

    est_rows = [
        (f"{year}년 1~{y2026['ytd_months']}월 누적 (방학월 제외)",
            fmt(y2026["sum_curr_ytd_ex"]) + "원"),
        (f"{next_year}년 1~{y2026['ytd_months']}월 누적 (방학월 제외)",
            fmt(y2026["sum_next_ex"]) + "원"),
        ("① YTD 실성장률  =  (위 두 값의 증감률)",
            yoy_str(y2026["ytd_yoy"])),
        (f"{year}년 연간 (방학월 제외)",
            fmt(annual_ex) + "원"),
        (f"② 1~{y2026['ytd_months']}월 비중 (당해/전년 평균)",
            f"{y2026['weight_pct']}%  (당해 {y2026['weight_curr']}% · 전년 {y2026['weight_prev']}%)"),
        (f"③ {next_year}년 연간 추정 (방학월 제외 환산)",
            fmt(y2026["est_annual_ex"]) + "원  (" + fmt_eok(y2026["est_annual_ex"]) + ")"),
        ("④ 추정 등락률  =  (③ vs " + f"{year}년 연간(방학월 제외))",
            yoy_str(y2026["est_yoy"])),
        (f"{next_year}년 연간 추정 (방학월 포함, 대시보드 KPI 표시값)",
            fmt(y2026["est_annual"]) + "원  (" + fmt_eok(y2026["est_annual"]) + ")"),
    ]
    for label, val in est_rows:
        lc = ws.cell(r, 2, label)
        lc.fill = P_GRAY; lc.font = FB; lc.alignment = AL; lc.border = BD
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        vc = ws.cell(r, 3, val)
        vc.font = FN; vc.alignment = AL; vc.border = BD
        for col in range(3, 7):
            ws.cell(r, col).border = BD
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:F{r}")
    src = ws.cell(r, 2,
        f"※ 등락률(%)은 방학월(1·7·12월) 매출 변동성 제거를 위해 방학월 제외 기준으로 산정합니다.  "
        f"|  출처: FoodnBid info.foodnbid.com  |  생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    src.font = FSRC; src.alignment = AL


@app.route("/api/excel")
def api_excel():
    """ANALYSIS_YEAR(2025) 엑셀 파일 우선 다운로드, 없으면 최신 파일.
    {next_year}년(진행중) 데이터가 있으면 최신 데이터로 진행현황 시트를 추가해 제공한다."""
    keyword = request.args.get("keyword", "").strip()
    if not keyword:
        return jsonify({"error": "keyword_required"}), 400

    def _files(pattern):
        return [f for f in sorted(glob.glob(pattern))
                if not os.path.basename(f).startswith("~$")]

    # 1순위: ANALYSIS_YEAR(2025) 파일
    preferred = _files(os.path.join(OUTPUT_DIR,
                       f"{keyword}_학교급식규모_{ANALYSIS_YEAR}_*.xlsx"))
    # 2순위: 전년도(2024) 파일
    fallback1 = _files(os.path.join(OUTPUT_DIR,
                       f"{keyword}_학교급식규모_{ANALYSIS_YEAR-1}_*.xlsx"))
    # 3순위: 연도 무관 최신 파일
    fallback2 = _files(os.path.join(OUTPUT_DIR,
                       f"{keyword}_학교급식규모_*.xlsx"))

    files = preferred or fallback1 or fallback2
    if not files:
        return jsonify({"error": "no_excel",
                        "message": "엑셀 파일이 없습니다. 스킬을 먼저 실행해 주세요."}), 404

    latest = files[-1]

    # {next_year}년(진행중) 데이터가 있으면 최신 데이터로 "진행현황" 시트를 추가한 임시 파일을 제공
    data = compute_data(keyword)
    if data and data.get("y2026"):
        try:
            import openpyxl, tempfile
            wb = openpyxl.load_workbook(latest)
            _add_y2026_sheet(wb, data)
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.close()
            wb.save(tmp.name)
            wb.close()

            @after_this_request
            def _cleanup(response):
                try:
                    os.remove(tmp.name)
                except OSError:
                    pass
                return response

            return send_file(
                tmp.name,
                as_attachment=True,
                download_name=os.path.basename(latest),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            print(f"[경고] {data['next_year']}년 진행현황 시트 추가 실패: {e}")

    return send_file(
        latest,
        as_attachment=True,
        download_name=os.path.basename(latest),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/status")
def api_status():
    keyword  = request.args.get("keyword", "").strip()
    status   = running_tasks.get(keyword, "idle")
    has_data = find_best_year(keyword) and get_latest_log(keyword, find_best_year(keyword)) is not None
    return jsonify({"status": status, "has_data": bool(has_data)})


# ── 실행 ──────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8765))
    is_local = not IS_CLOUD
    if is_local:
        import webbrowser
        threading.Timer(1.5, lambda: webbrowser.open(f"http://localhost:{port}")).start()
    url = f"http://localhost:{port}"
    print("=" * 55)
    print("  학교급식 시장규모 대시보드")
    print(f"  접속 주소: {url}")
    print("  종료: Ctrl+C")
    print("=" * 55)
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
