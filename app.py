"""
식품 시장 분석 대시보드
푸드앤비드 수도권 매출 + 블루시스마켓 상품 데이터 통합 분석
"""

import asyncio
import logging
import os
import sys
import threading
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# 프로젝트 루트를 sys.path에 추가
ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

import io
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from exporters.excel_exporter import create_excel_report
from processors.data_analyzer import DataAnalyzer
from scrapers.bluesys_scraper import BluesysMarketScraper
from scrapers.foodnbid_scraper import FoodNBidScraper
from scrapers.foodspring_scraper import FoodspringScraper

# ============================================================
# 로깅 설정
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ============================================================
# 페이지 설정
# ============================================================
st.set_page_config(
    page_title="식품 시장 분석 대시보드",
    page_icon="🍖",
    layout="wide",
    initial_sidebar_state="collapsed",
)
# 사이드바 완전 숨김 + 전체 레이아웃 조정
st.markdown("""
<style>
  [data-testid="stSidebar"] { display: none !important; }
  [data-testid="collapsedControl"] { display: none !important; }
  .block-container { padding-top: 1rem !important; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# 커스텀 CSS
# ============================================================
st.markdown("""
<style>
    .main-title {
        font-size: 2rem;
        font-weight: 800;
        color: #1F4E79;
        margin-bottom: 0.2rem;
    }
    .sub-title {
        font-size: 1rem;
        color: #595959;
        margin-bottom: 1.5rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #1F4E79, #2E86AB);
        color: white;
        border-radius: 12px;
        padding: 1rem 1.5rem;
        text-align: center;
    }
    .metric-label { font-size: 0.85rem; opacity: 0.85; }
    .metric-value { font-size: 1.6rem; font-weight: 800; }
    .insight-box {
        background: #FFFBEA;
        border-left: 5px solid #F4A300;
        border-radius: 4px;
        padding: 1rem 1.5rem;
        margin: 1rem 0;
        font-size: 0.95rem;
    }
    .stProgress > div > div > div { background-color: #1F4E79; }
    div[data-testid="stSidebarContent"] { background: #F0F4FA; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# 세션 상태 초기화
# ============================================================
def init_session():
    defaults = {
        "sales_df": pd.DataFrame(),
        "product_df": pd.DataFrame(),
        "analysis_done": False,
        "competitive": {},
        "sales_type_df": pd.DataFrame(),
        "top_products_df": pd.DataFrame(),
        "regional_df": pd.DataFrame(),
        "excel_bytes": None,
        "last_keyword": "",
        "last_dates": ("", ""),
        # 식봄 전용
        "foodspring_df": pd.DataFrame(),
        "foodspring_excel_bytes": None,
        "foodspring_done": False,
        # 품목제조번호 조회 전용
        "prdlst_api_key": "",
        "prdlst_results": [],
        "prdlst_excel_bytes": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_session()


# ============================================================
# 사이드바 - 숨김 처리 (변수 기본값만 유지)
# ============================================================
fnb_id = fnb_pw = bls_id = bls_pw = ""
keyword = ""
start_date = date.today() - timedelta(days=365)
end_date = date.today()
max_products = 50
headless_mode = True
run_btn = False

with st.sidebar:
    st.markdown("## ⚙️ 설정")
    st.divider()

    # --- 로그인 정보 ---
    st.markdown("### 🔐 로그인 정보")
    with st.expander("푸드앤비드 계정", expanded=True):
        fnb_id = st.text_input("아이디", key="fnb_id", placeholder="아이디 입력")
        fnb_pw = st.text_input("비밀번호", key="fnb_pw", type="password", placeholder="비밀번호 입력")

    with st.expander("블루시스마켓 계정", expanded=True):
        bls_id = st.text_input("아이디", key="bls_id", placeholder="아이디 입력")
        bls_pw = st.text_input("비밀번호", key="bls_pw", type="password", placeholder="비밀번호 입력")

    st.divider()

    # --- 검색 조건 ---
    st.markdown("### 🔍 검색 조건")
    keyword = st.text_input(
        "제품 키워드",
        placeholder="예: 돈까스, 소시지, 비엔나, 후랑크",
        help="검색할 제품의 대표 키워드를 입력하세요"
    )

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input(
            "시작일",
            value=date.today() - timedelta(days=365),
            format="YYYY-MM-DD",
        )
    with col2:
        end_date = st.date_input(
            "종료일",
            value=date.today(),
            format="YYYY-MM-DD",
        )

    max_products = st.slider(
        "블루시스마켓 최대 수집 상품 수",
        min_value=10,
        max_value=200,
        value=50,
        step=10,
        help="많을수록 정확하지만 시간이 오래 걸립니다"
    )

    headless_mode = st.toggle(
        "백그라운드 실행 (헤드리스)",
        value=True,
        help="OFF 시 브라우저 창이 보입니다 (디버깅용)"
    )

    st.divider()

    # --- 실행 버튼 ---
    run_btn = st.button(
        "🚀 데이터 수집 & 분석 시작",
        type="primary",
        use_container_width=True,
        disabled=not (keyword and fnb_id and fnb_pw and bls_id and bls_pw),
    )

    if not keyword:
        st.caption("⚠️ 키워드를 입력해주세요")
    if not (fnb_id and fnb_pw):
        st.caption("⚠️ 푸드앤비드 계정을 입력해주세요")
    if not (bls_id and bls_pw):
        st.caption("⚠️ 블루시스마켓 계정을 입력해주세요")


# ============================================================
# 메인 콘텐츠
# ============================================================
MANUAL_URL = "https://raw.githubusercontent.com/petitfarming-lab/food-market-analyzer/main/manual.pdf"

# ── 방문자 카운터 (세션당 1회, 파일 기반) ──
import json, threading
_COUNTER_FILE = "/tmp/visit_counts.json"
_counter_lock = threading.Lock()

def _load_counts():
    try:
        with open(_COUNTER_FILE) as f:
            return json.load(f)
    except Exception:
        return {"total": 0, "daily": {}}

def _save_counts(c):
    try:
        with open(_COUNTER_FILE, "w") as f:
            json.dump(c, f)
    except Exception:
        pass

def _increment_visit():
    today = date.today().strftime("%Y-%m-%d")
    with _counter_lock:
        c = _load_counts()
        c["total"] = c.get("total", 0) + 1
        c.setdefault("daily", {})
        c["daily"][today] = c["daily"].get(today, 0) + 1
        _save_counts(c)
        return c["daily"][today], c["total"]

if not st.session_state.get("_visited"):
    st.session_state["_visited"] = True
    _t, _tot = _increment_visit()
    st.session_state["_today_cnt"] = f"{_t:,}"
    st.session_state["_total_cnt"] = f"{_tot:,}"

_today_cnt = st.session_state.get("_today_cnt", "—")
_total_cnt = st.session_state.get("_total_cnt", "—")

st.markdown(f"""
<div style="display:flex; justify-content:space-between; align-items:center;
            background:#1B3A6B; padding:10px 20px; border-radius:8px; margin-bottom:12px;">
  <span style="color:white; font-size:0.82rem; letter-spacing:0.04em;">
    © 2026 CJ CheilJedang B2B Marketing
    &nbsp;&nbsp;|&nbsp;&nbsp;
    <span style="color:#90C4F5;">오늘 접속 {_today_cnt}회</span>
    &nbsp;·&nbsp;
    <span style="color:#90C4F5;">누계 {_total_cnt}회</span>
  </span>
  <a href="{MANUAL_URL}" target="_blank"
     style="color:#90C4F5; font-size:0.82rem; text-decoration:none; font-weight:600;
            border:1px solid #4A7BC4; padding:4px 12px; border-radius:4px;">
    📄 사용 설명서 다운로드
  </a>
</div>
<div class="main-title">🍖 식품 분석 대시보드</div>
""", unsafe_allow_html=True)

# ============================================================
# 메인 탭 분리
# ============================================================
main_tab2, main_tab3, main_tab4, main_tab5, main_tab1 = st.tabs([
    "🛒 식봄 가격 조회",
    "🔍 품목제조번호 조회",
    "📊 생산실적 조회",
    "📋 월별예상금액 수집",
    "　",
])


# ============================================================
# 데이터 수집 실행 (공통 함수)
# ============================================================
async def run_scraping(keyword, start_date, end_date, max_products, headless):
    """비동기 스크래핑 실행"""
    results = {"sales_df": pd.DataFrame(), "product_df": pd.DataFrame()}

    progress_placeholder = st.empty()
    status_placeholder = st.empty()

    def update_progress(msg: str):
        status_placeholder.info(f"⏳ {msg}")

    # 푸드앤비드 수집
    try:
        update_progress("푸드앤비드 접속 중...")
        async with FoodNBidScraper(fnb_id, fnb_pw, headless=headless) as scraper:
            results["sales_df"] = await scraper.get_sales_data(
                keyword=keyword,
                start_date=str(start_date),
                end_date=str(end_date),
                progress_callback=update_progress,
            )
    except Exception as e:
        st.warning(f"⚠️ 푸드앤비드 수집 오류: {e}\n스크린샷: debug_foodnbid_error.png 확인")
        logger.error(f"FoodNBid 오류: {e}")

    # 블루시스마켓 수집
    try:
        update_progress("블루시스마켓 접속 중...")
        async with BluesysMarketScraper(bls_id, bls_pw, headless=headless) as scraper:
            results["product_df"] = await scraper.get_product_data(
                keyword=keyword,
                max_products=max_products,
                progress_callback=update_progress,
            )
    except Exception as e:
        st.warning(f"⚠️ 블루시스마켓 수집 오류: {e}\n스크린샷: debug_bluesys_error.png 확인")
        logger.error(f"Bluesys 오류: {e}")

    status_placeholder.success("✅ 데이터 수집 완료!")
    return results


# ============================================================
# 식봄 스크래핑 함수
# ============================================================
async def run_foodspring_scraping(df_input: pd.DataFrame, headless: bool):
    result_holder = {"df": pd.DataFrame()}
    status = st.empty()

    def update_progress(msg: str):
        status.info(f"⏳ {msg}")

    try:
        async with FoodspringScraper(headless=headless) as scraper:
            result_holder["df"] = await scraper.search_products(
                df_input=df_input,
                progress_callback=update_progress,
            )
        status.success("✅ 식봄 검색 완료!")
    except Exception as e:
        logger.error(f"식봄 스크래핑 오류: {e}")
        status.error(f"오류 발생: {e}")

    return result_holder["df"]


def build_foodspring_excel(df: pd.DataFrame) -> bytes:
    """식봄 검색 결과 엑셀 생성 후 bytes 반환"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "식봄_검색결과"

    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2F75B6")
    cell_font = Font(name="맑은 고딕", size=10)
    found_fill = PatternFill("solid", start_color="E8F4FD")
    notfound_fill = PatternFill("solid", start_color="FFF2CC")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["No.", "입력_제품명", "입력_제조사", "순위", "식봄_판매제품명", "식봄_판매가", "상품URL", "검색상태"]
    col_widths = [5, 28, 18, 6, 42, 14, 50, 10]

    for c_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.row_dimensions[1].height = 22

    for _, row in df.iterrows():
        r_idx = ws.max_row + 1
        status = row.get("검색상태", "")
        row_fill = found_fill if status == "발견" else notfound_fill
        data = [
            row.get("No", ""),
            row.get("제품명", ""),
            row.get("제조사", ""),
            row.get("순위", "-"),
            row.get("판매제품명", ""),
            row.get("판매가", "-"),
            row.get("상품URL", ""),
            status,
        ]
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.fill = row_fill
            cell.border = border
            cell.alignment = center if c_idx in (1, 4, 6, 8) else left

    # 같은 No. 묶음 병합
    if not df.empty and "No" in df.columns:
        groups = df.groupby("No", sort=False)
        for no, grp in groups:
            if len(grp) > 1:
                start_r = grp.index[0] + 2
                end_r = grp.index[-1] + 2
                for mc in (1, 2, 3):
                    ws.merge_cells(start_row=start_r, start_column=mc, end_row=end_r, end_column=mc)
                    ws.cell(start_r, mc).alignment = center

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================
# main_tab1: 기존 시장 분석 UI
# ============================================================
with main_tab1:
    if run_btn:
        with st.spinner("데이터 수집 중... (수분 소요될 수 있습니다)"):
            result_container = {}

            def scrape_in_thread():
                if sys.platform == "win32":
                    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    result_container["results"] = loop.run_until_complete(
                        run_scraping(keyword, start_date, end_date, max_products, headless_mode)
                    )
                except Exception as e:
                    logger.error(f"스크래핑 스레드 오류: {e}")
                    result_container["results"] = {
                        "sales_df": pd.DataFrame(),
                        "product_df": pd.DataFrame(),
                    }
                finally:
                    loop.close()

            t = threading.Thread(target=scrape_in_thread, daemon=True)
            t.start()
            t.join()

            results = result_container.get(
                "results",
                {"sales_df": pd.DataFrame(), "product_df": pd.DataFrame()}
            )

        st.session_state["sales_df"] = results["sales_df"]
        st.session_state["product_df"] = results["product_df"]
        st.session_state["last_keyword"] = keyword
        st.session_state["last_dates"] = (str(start_date), str(end_date))

        analyzer = DataAnalyzer()
        if not results["sales_df"].empty:
            st.session_state["sales_type_df"] = analyzer.analyze_sales_by_type(results["sales_df"])
            st.session_state["top_products_df"] = analyzer.get_top_products(results["sales_df"])
            st.session_state["regional_df"] = analyzer.get_regional_breakdown(results["sales_df"])
        else:
            st.session_state["sales_type_df"] = pd.DataFrame()
            st.session_state["top_products_df"] = pd.DataFrame()
            st.session_state["regional_df"] = pd.DataFrame()

        st.session_state["competitive"] = analyzer.analyze_competitive_spec(
            results["product_df"], results["sales_df"]
        )

        try:
            excel_bytes = create_excel_report(
                sales_df=results["sales_df"],
                product_df=results["product_df"],
                sales_type_df=st.session_state["sales_type_df"],
                top_products_df=st.session_state["top_products_df"],
                regional_df=st.session_state["regional_df"],
                competitive_analysis=st.session_state["competitive"],
                keyword=keyword,
                start_date=str(start_date),
                end_date=str(end_date),
            )
            st.session_state["excel_bytes"] = excel_bytes
        except Exception as e:
            st.warning(f"엑셀 생성 오류: {e}")
            logger.error(f"Excel 오류: {e}")

        st.session_state["analysis_done"] = True
        st.rerun()

    if st.session_state["analysis_done"]:
        kw = st.session_state["last_keyword"]
        s_date, e_date = st.session_state["last_dates"]
        sales_df = st.session_state["sales_df"]
        product_df = st.session_state["product_df"]
        sales_type_df = st.session_state["sales_type_df"]
        top_products_df = st.session_state["top_products_df"]
        regional_df = st.session_state["regional_df"]
        competitive = st.session_state["competitive"]

        st.info(f"**분석 대상**: [{kw}] | **기간**: {s_date} ~ {e_date} | **수도권 기준**")

        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        total_sales = int(sales_df["매출금액"].sum()) if not sales_df.empty and "매출금액" in sales_df.columns else 0
        total_qty = int(sales_df["수량"].sum()) if not sales_df.empty and "수량" in sales_df.columns else 0
        total_products = len(product_df) if not product_df.empty else 0
        median_price = competitive.get("median_price", 0)

        with kpi1:
            st.metric("💰 수도권 총매출", f"{total_sales:,}원" if total_sales else "수집중")
        with kpi2:
            st.metric("📦 총 판매 수량", f"{total_qty:,}개" if total_qty else "수집중")
        with kpi3:
            st.metric("🛒 분석 상품 수", f"{total_products:,}개")
        with kpi4:
            st.metric("💲 학교매입가 중앙값", f"{median_price:,}원" if median_price else "-")

        st.divider()

        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📊 매출 타입 분석", "🏆 TOP 제품", "🗺 지역별 분석", "🛒 상품 상세", "💡 경쟁력 분석",
        ])

        with tab1:
            if not sales_type_df.empty:
                col_chart, col_table = st.columns([1, 1])
                with col_chart:
                    st.subheader("제품 타입별 매출 비중")
                    fig_pie = px.pie(
                        sales_type_df, names="제품타입", values="총매출금액",
                        color_discrete_sequence=px.colors.qualitative.Set2, hole=0.4,
                    )
                    fig_pie.update_traces(textposition="inside", textinfo="percent+label")
                    fig_pie.update_layout(height=380, margin=dict(t=20, b=20))
                    st.plotly_chart(fig_pie, use_container_width=True)
                with col_table:
                    st.subheader("제품 타입별 매출금액")
                    fig_bar = px.bar(
                        sales_type_df.sort_values("총매출금액"), x="총매출금액", y="제품타입",
                        orientation="h", color="총매출금액", color_continuous_scale="Blues", text="총매출금액",
                    )
                    fig_bar.update_traces(texttemplate="%{text:,.0f}원", textposition="outside")
                    fig_bar.update_layout(height=380, margin=dict(t=20, b=20), showlegend=False)
                    st.plotly_chart(fig_bar, use_container_width=True)
                st.subheader("타입별 상세 수치")
                st.dataframe(
                    sales_type_df.style.format({
                        "총매출금액": "{:,}원", "총수량": "{:,}개",
                        "평균단가": "{:,}원", "매출비중(%)": "{:.2f}%",
                    }),
                    use_container_width=True,
                )
            else:
                st.info("푸드앤비드 매출 데이터를 수집하면 타입별 분석이 표시됩니다.")

        with tab2:
            if not top_products_df.empty:
                st.subheader(f"매출 상위 제품 TOP {len(top_products_df)}")
                fig_top = px.bar(
                    top_products_df.head(15).sort_values("총매출금액"), x="총매출금액", y="제품명",
                    orientation="h", color="평균단가", color_continuous_scale="RdYlGn_r",
                    title="매출 TOP 15 제품 (색상: 평균단가)",
                )
                fig_top.update_layout(height=480, margin=dict(t=40, b=20))
                st.plotly_chart(fig_top, use_container_width=True)
                st.dataframe(
                    top_products_df.style.format({
                        "총매출금액": "{:,}원", "총수량": "{:,}개", "평균단가": "{:,.0f}원",
                    }),
                    use_container_width=True,
                )
            else:
                st.info("푸드앤비드 매출 데이터를 수집하면 TOP 제품이 표시됩니다.")

        with tab3:
            if not regional_df.empty:
                col_map, col_tbl = st.columns([1, 1])
                with col_map:
                    st.subheader("지역별 매출금액")
                    fig_region = px.bar(
                        regional_df, x="지역", y="총매출금액", color="총매출금액",
                        color_continuous_scale="Blues", text="총매출금액",
                    )
                    fig_region.update_traces(texttemplate="%{text:,.0f}원", textposition="outside")
                    fig_region.update_layout(height=360, margin=dict(t=20, b=20))
                    st.plotly_chart(fig_region, use_container_width=True)
                with col_tbl:
                    st.subheader("지역별 상세")
                    st.dataframe(
                        regional_df.style.format({"총매출금액": "{:,}원", "총수량": "{:,}개"}),
                        use_container_width=True,
                    )
            else:
                st.info("지역 데이터가 없거나 수집이 필요합니다.")

        with tab4:
            if not product_df.empty:
                st.subheader(f"블루시스마켓 상품 목록 ({len(product_df)}개)")
                f_col1, f_col2, f_col3 = st.columns(3)
                with f_col1:
                    tier_filter = st.multiselect(
                        "TIER 필터",
                        options=product_df["TIER"].dropna().unique().tolist() if "TIER" in product_df.columns else [],
                    )
                with f_col2:
                    cook_filter = st.multiselect(
                        "조리법 필터",
                        options=product_df["조리법"].dropna().unique().tolist() if "조리법" in product_df.columns else [],
                    )
                with f_col3:
                    if "학교매입가_숫자" in product_df.columns:
                        price_range = st.slider(
                            "학교매입가 범위", min_value=0,
                            max_value=int(product_df["학교매입가_숫자"].max() or 50000),
                            value=(0, int(product_df["학교매입가_숫자"].max() or 50000)),
                        )
                    else:
                        price_range = (0, 999999)
                filtered_df = product_df.copy()
                if tier_filter and "TIER" in filtered_df.columns:
                    filtered_df = filtered_df[filtered_df["TIER"].isin(tier_filter)]
                if cook_filter and "조리법" in filtered_df.columns:
                    filtered_df = filtered_df[filtered_df["조리법"].isin(cook_filter)]
                if "학교매입가_숫자" in filtered_df.columns:
                    filtered_df = filtered_df[filtered_df["학교매입가_숫자"].between(price_range[0], price_range[1])]
                display_cols = ["제품명", "학교매입가", "함량", "TIER", "조리법", "제조사", "브랜드"]
                disp_cols = [c for c in display_cols if c in filtered_df.columns]
                st.dataframe(filtered_df[disp_cols], use_container_width=True, height=400)
                if "학교매입가_숫자" in filtered_df.columns:
                    price_data = filtered_df["학교매입가_숫자"].replace(0, pd.NA).dropna()
                    if not price_data.empty:
                        fig_hist = px.histogram(
                            price_data, nbins=20, title="학교매입가 분포",
                            color_discrete_sequence=["#1F4E79"],
                        )
                        fig_hist.update_layout(height=280, margin=dict(t=40, b=20))
                        st.plotly_chart(fig_hist, use_container_width=True)
            else:
                st.info("블루시스마켓 상품 데이터를 수집하면 표시됩니다.")

        with tab5:
            insight_text = competitive.get("insight_text", "")
            if insight_text:
                st.markdown(f'<div class="insight-box">{insight_text}</div>', unsafe_allow_html=True)
            col_a, col_b = st.columns(2)
            with col_a:
                tier_dist = competitive.get("tier_distribution", {})
                if tier_dist:
                    st.subheader("TIER 분포")
                    fig_tier = px.pie(
                        names=list(tier_dist.keys()), values=list(tier_dist.values()),
                        color_discrete_sequence=px.colors.qualitative.Pastel, hole=0.35,
                    )
                    fig_tier.update_layout(height=300, margin=dict(t=10, b=10))
                    st.plotly_chart(fig_tier, use_container_width=True)
            with col_b:
                cook_dist = competitive.get("cooking_method_distribution", {})
                if cook_dist:
                    st.subheader("조리법 분포")
                    fig_cook = px.bar(
                        x=list(cook_dist.keys()), y=list(cook_dist.values()),
                        color_discrete_sequence=["#2E86AB"],
                    )
                    fig_cook.update_layout(height=300, margin=dict(t=10, b=10))
                    st.plotly_chart(fig_cook, use_container_width=True)
            comp_products = competitive.get("competitive_products", pd.DataFrame())
            if not comp_products.empty:
                st.subheader("매출 TOP & 상품 스펙 교차 분석")
                st.dataframe(comp_products, use_container_width=True)

        st.divider()
        dl_col1, dl_col2, dl_col3 = st.columns([1, 2, 1])
        with dl_col2:
            if st.session_state.get("excel_bytes"):
                filename = f"식품시장분석_{kw}_{s_date}_{e_date}.xlsx"
                st.download_button(
                    label="📥 엑셀 보고서 다운로드",
                    data=st.session_state["excel_bytes"],
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
                st.caption(f"파일명: {filename}")
    else:
        pass


# ============================================================
# main_tab2: 식봄 가격 조회 UI
# ============================================================
with main_tab2:
    st.markdown("### 🛒 식봄(foodspring.co.kr) 제품 가격 조회")
    st.info("아래 표에 **엑셀에서 복사(Ctrl+C)한 데이터를 바로 붙여넣기(Ctrl+V)** 하세요. 직접 입력도 가능합니다.")

    # ── 입력 테이블 초기화 ──
    if "fs_input_df" not in st.session_state:
        st.session_state["fs_input_df"] = pd.DataFrame(
            {"제품명": [""] * 10, "제조사": [""] * 10}
        )

    fs_top_col1, fs_top_col2 = st.columns([4, 1])
    with fs_top_col2:
        if st.button("🗑️ 초기화", key="fs_clear", use_container_width=True):
            st.session_state["fs_input_df"] = pd.DataFrame(
                {"제품명": [""] * 10, "제조사": [""] * 10}
            )
            st.session_state["foodspring_done"] = False
            st.session_state["foodspring_df"] = pd.DataFrame()
            st.rerun()
        fs_headless = st.toggle("백그라운드 실행", value=True, key="fs_headless",
                                help="OFF 시 브라우저 창이 보입니다")

    with fs_top_col1:
        st.caption("📋 엑셀에서 복사 → 첫 번째 셀 클릭 → Ctrl+V 로 붙여넣기 | 행 추가: 표 하단 + 버튼")

    # ── data_editor: 붙여넣기/직접입력 가능한 표 ──
    edited_df = st.data_editor(
        st.session_state["fs_input_df"],
        num_rows="dynamic",
        use_container_width=True,
        height=320,
        key="fs_editor",
        column_config={
            "제품명": st.column_config.TextColumn("제품명 ✱ (필수)", width="large"),
            "제조사": st.column_config.TextColumn("제조사 (선택)", width="medium"),
        },
    )

    # 유효 행만 추출
    df_input = edited_df.copy()
    df_input.columns = df_input.columns.str.strip()
    df_input = df_input.fillna("").astype(str)
    df_input = df_input[df_input["제품명"].str.strip() != ""].reset_index(drop=True)

    valid_count = len(df_input)
    st.caption(f"유효 제품 수: **{valid_count}개**")

    fs_run_btn = st.button(
        f"🔍 식봄 가격 검색 시작  ({valid_count}개)",
        type="primary",
        disabled=(valid_count == 0),
        key="fs_run",
    )

    if fs_run_btn:
        # 현재 입력 저장
        st.session_state["fs_input_df"] = edited_df
        st.session_state["foodspring_done"] = False

        with st.spinner("식봄에서 제품 검색 중... 잠시 기다려 주세요."):
            fs_result_container = {}

            def fs_scrape_thread():
                if sys.platform == "win32":
                    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    fs_result_container["df"] = loop.run_until_complete(
                        run_foodspring_scraping(df_input, fs_headless)
                    )
                except Exception as e:
                    logger.error(f"식봄 스레드 오류: {e}")
                    fs_result_container["df"] = pd.DataFrame()
                finally:
                    loop.close()

            fs_t = threading.Thread(target=fs_scrape_thread, daemon=True)
            fs_t.start()
            fs_t.join()

        fs_df = fs_result_container.get("df", pd.DataFrame())
        st.session_state["foodspring_df"] = fs_df
        st.session_state["foodspring_done"] = True

        if not fs_df.empty:
            try:
                st.session_state["foodspring_excel_bytes"] = build_foodspring_excel(fs_df)
            except Exception as e:
                logger.error(f"식봄 엑셀 생성 오류: {e}")

        st.rerun()

    # ── 결과 표시 ──
    if st.session_state.get("foodspring_done") and not st.session_state["foodspring_df"].empty:
        fs_df = st.session_state["foodspring_df"]
        total = fs_df["No"].nunique() if "No" in fs_df.columns else len(fs_df)
        found = fs_df[fs_df["검색상태"] == "발견"]["No"].nunique() if "검색상태" in fs_df.columns else 0

        st.divider()
        st.markdown("#### 검색 결과")

        m1, m2, m3 = st.columns(3)
        m1.metric("총 검색 제품", f"{total}개")
        m2.metric("✅ 발견", f"{found}개")
        m3.metric("❌ 미발견", f"{total - found}개")

        st.divider()

        display_cols = ["No", "제품명", "제조사", "순위", "판매제품명", "판매가", "검색상태"]
        show_cols = [c for c in display_cols if c in fs_df.columns]

        def highlight_status(row):
            if row.get("검색상태") == "미발견":
                return ["background-color: #FFF9C4"] * len(row)
            return [""] * len(row)

        st.dataframe(
            fs_df[show_cols].style.apply(highlight_status, axis=1),
            use_container_width=True,
            height=450,
        )

        if st.session_state.get("foodspring_excel_bytes"):
            from datetime import datetime as _dt
            fname = f"식봄_가격조회_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
            dl_c1, dl_c2, dl_c3 = st.columns([1, 2, 1])
            with dl_c2:
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=st.session_state["foodspring_excel_bytes"],
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )


# ============================================================
# main_tab3: 품목제조번호 조회
# ============================================================
PRDLST_API_BASE = "https://openapi.foodsafetykorea.go.kr/api"

PRDLST_SERVICE_MAP = {
    "C006":  "축산물",
    "I1250": "식품/식품첨가물",
    "I0030": "건강기능식품",
}

PRDLST_FIELD_KO = {
    "PRDLST_REPORT_NO": "품목보고번호", "PRMS_DT": "일자", "PRDLST_NM": "제품명",
    "INDUTY_CD_NM": "유형", "POG_DAYCNT": "소비기한", "DISPOS": "제품형태",
    "PACKING_UNIT_DESCRIP": "포장재질", "RAWMTRL_NM": "원재료명 및 함량",
    "HIENG_LNTRT_DVS_NM": "내수/수출/겸용", "BSSH_NM": "업체명",
    "SITE_ADDR": "소재지", "LCNS_NO": "허가번호", "CHNG_DT": "최종변경일",
    "LAST_UPDT_DTM": "최종업데이트", "INDUTY_CD": "업종코드",
    "PRDLST_DCNM": "식품유형(상세)", "PRODUCTION": "생산여부",
    "USAGE": "용도", "RAW_MTRL_NM": "원재료명 및 함량",
    "RAWMTRL_ORDNO": "원재료순번",
}


def prdlst_fetch_one(no: str, svc_code: str, api_key: str):
    """품목제조번호 1건 조회. 성공 시 dict 반환, 실패 시 예외."""
    import urllib.request, urllib.parse, json as _json
    url = f"{PRDLST_API_BASE}/{api_key}/{svc_code}/json/1/200/PRDLST_REPORT_NO={urllib.parse.quote(no)}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=10) as r:
        body = r.read()
    if body.lstrip().startswith(b"<"):
        raise ValueError("API 키가 유효하지 않거나 해당 서비스 권한이 없습니다.")
    data = _json.loads(body)
    svc_data = data.get(svc_code)
    if not svc_data:
        raise ValueError("응답 형식 오류")
    code = svc_data.get("RESULT", {}).get("CODE", "")
    if code == "INFO-200" or not svc_data.get("row"):
        raise ValueError("조회 결과 없음")
    if code.startswith("ERROR"):
        raise ValueError(f"API 오류: {code}")
    rows = svc_data["row"]
    item = rows[0]

    # 원재료 순서 처리
    ingredients = []
    if len(rows) > 1:
        paired = [
            {"name": (r.get("RAWMTRL_NM") or r.get("RAW_MTRL_NM") or "").strip(),
             "order": int((r.get("RAWMTRL_ORDNO") or "9999").strip()) if str(r.get("RAWMTRL_ORDNO") or "9999").strip().isdigit() else 9999}
            for r in rows if (r.get("RAWMTRL_NM") or r.get("RAW_MTRL_NM") or "").strip()
        ]
        paired.sort(key=lambda x: x["order"])
        ingredients = [p["name"] for p in paired]
    else:
        raw = (item.get("RAWMTRL_NM") or item.get("RAW_MTRL_NM") or "").strip()
        ordno_str = (item.get("RAWMTRL_ORDNO") or "").strip()
        if raw:
            names = [s.strip() for s in raw.split(",") if s.strip()]
            if ordno_str:
                ordnos = []
                for s in ordno_str.split(","):
                    s = s.strip()
                    ordnos.append(int(s) if s.isdigit() else 9999)
                paired2 = sorted(zip(names, ordnos), key=lambda x: x[1])
                ingredients = [n for n, _ in paired2]
            else:
                ingredients = names

    item["_ingredients"] = ingredients
    item["_svc_code"] = svc_code
    return item


def build_prdlst_excel(results: list) -> bytes:
    """품목제조번호 조회 결과를 엑셀로 변환. 원재료 열별 분리."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "품목제조번호조회"

    header_font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    section_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    section_fill = PatternFill("solid", start_color="2E86AB")
    cell_font    = Font(name="맑은 고딕", size=10)
    alt_fill     = PatternFill("solid", start_color="EBF5FB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 최대 원재료 수 계산
    max_ing = max((len(r["item"].get("_ingredients", [])) for r in results if r.get("item")), default=0)

    base_headers = [
        "품목보고번호", "제품명", "유형", "업체명", "소재지",
        "일자", "소비기한", "제품형태", "포장재질", "내수/수출/겸용",
        "허가번호", "최종변경일", "식품유형(상세)", "생산여부", "용도",
    ]
    ing_headers = [f"원재료{i+1}" for i in range(max_ing)]
    all_headers = base_headers + ing_headers + ["식품유형"]

    for c_idx, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(c_idx)].width = 18 if c_idx <= len(base_headers) else 22

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 30
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    field_map = {
        "품목보고번호": "PRDLST_REPORT_NO", "제품명": "PRDLST_NM",
        "유형": "INDUTY_CD_NM", "업체명": "BSSH_NM", "소재지": "SITE_ADDR",
        "일자": "PRMS_DT", "소비기한": "POG_DAYCNT", "제품형태": "DISPOS",
        "포장재질": "PACKING_UNIT_DESCRIP", "내수/수출/겸용": "HIENG_LNTRT_DVS_NM",
        "허가번호": "LCNS_NO", "최종변경일": "CHNG_DT",
        "식품유형(상세)": "PRDLST_DCNM", "생산여부": "PRODUCTION", "용도": "USAGE",
    }

    for r_idx, res in enumerate(results, 2):
        item = res.get("item")
        row_fill = alt_fill if r_idx % 2 == 0 else None
        if not item:
            cell = ws.cell(row=r_idx, column=1, value=res.get("no", ""))
            cell.font = cell_font
            cell.border = border
            err_cell = ws.cell(row=r_idx, column=2, value=f"조회 실패: {res.get('error','')}")
            err_cell.font = Font(name="맑은 고딕", size=10, color="CC0000")
            err_cell.border = border
            continue

        for c_idx, h in enumerate(base_headers, 1):
            val = item.get(field_map.get(h, ""), "") or ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.alignment = left
            cell.border = border
            if row_fill:
                cell.fill = row_fill

        # 원재료 열별 입력
        ingredients = item.get("_ingredients", [])
        for i, ing in enumerate(ingredients):
            c_idx = len(base_headers) + i + 1
            cell = ws.cell(row=r_idx, column=c_idx, value=ing)
            cell.font = cell_font
            cell.alignment = left
            cell.border = border
            if row_fill:
                cell.fill = row_fill

        # 식품유형(서비스코드)
        svc = item.get("_svc_code", "")
        last_col = len(base_headers) + max_ing + 1
        cell = ws.cell(row=r_idx, column=last_col, value=PRDLST_SERVICE_MAP.get(svc, svc))
        cell.font = cell_font
        cell.alignment = center
        cell.border = border
        if row_fill:
            cell.fill = row_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


with main_tab3:
    from datetime import datetime as _dt2

    st.markdown("### 🔍 품목제조번호 조회 (식품안전나라)")
    st.info("품목제조번호를 한 줄에 하나씩 입력하고 식품 유형을 선택 후 조회하세요. 최대 50개 일괄 조회 가능합니다.")

    # ── API 키 설정 ──
    with st.expander("🔑 API 인증키 설정", expanded=(not st.session_state.get("prdlst_api_key"))):
        api_key_input = st.text_input(
            "식품안전나라 OpenAPI 인증키",
            value=st.session_state.get("prdlst_api_key", ""),
            type="password",
            placeholder="발급받은 인증키 입력",
            key="prdlst_api_key_input",
        )
        col_save, col_link = st.columns([1, 2])
        with col_save:
            if st.button("저장", key="prdlst_save_key"):
                st.session_state["prdlst_api_key"] = api_key_input.strip()
                st.success("인증키 저장됨")
        with col_link:
            st.markdown("[식품안전나라 OpenAPI 신청 →](https://www.foodsafetykorea.go.kr/api/main.do)")

    # ── 식품 유형 선택 ──
    svc_label_map = {v: k for k, v in PRDLST_SERVICE_MAP.items()}
    svc_selected_label = st.radio(
        "식품 유형",
        options=list(PRDLST_SERVICE_MAP.values()),
        horizontal=True,
        key="prdlst_svc_label",
    )
    svc_code_selected = svc_label_map[svc_selected_label]

    # ── 번호 입력 ──
    numbers_raw = st.text_area(
        "품목제조번호 (한 줄에 하나)",
        height=160,
        placeholder="예:\n2013047600422\n2015040500139",
        key="prdlst_numbers",
    )

    nos = [n.strip() for n in numbers_raw.strip().splitlines() if n.strip()]
    st.caption(f"입력된 번호: **{len(nos)}개** (최대 50개)")

    p3_run = st.button(
        f"🔍 조회하기 ({len(nos)}개)",
        type="primary",
        disabled=(len(nos) == 0 or not st.session_state.get("prdlst_api_key")),
        key="prdlst_run",
    )
    if not st.session_state.get("prdlst_api_key"):
        st.caption("⚠️ 위에서 API 인증키를 먼저 저장하세요.")

    if p3_run:
        key = st.session_state["prdlst_api_key"]
        nos_limited = nos[:50]
        results_list = []
        prog = st.progress(0)
        status_ph = st.empty()
        for i, no in enumerate(nos_limited):
            status_ph.info(f"조회 중 [{i+1}/{len(nos_limited)}]: {no}")
            try:
                item = prdlst_fetch_one(no, svc_code_selected, key)
                results_list.append({"no": no, "item": item, "error": None})
            except Exception as e:
                results_list.append({"no": no, "item": None, "error": str(e)})
            prog.progress((i + 1) / len(nos_limited))
        status_ph.success(f"✅ 조회 완료 ({len(nos_limited)}개)")
        st.session_state["prdlst_results"] = results_list

        ok_list = [r for r in results_list if r["item"]]
        if ok_list:
            try:
                st.session_state["prdlst_excel_bytes"] = build_prdlst_excel(results_list)
            except Exception as e:
                logger.error(f"품목제조 엑셀 오류: {e}")
        st.rerun()

    # ── 결과 표시 ──
    results_list = st.session_state.get("prdlst_results", [])
    if results_list:
        ok_cnt  = sum(1 for r in results_list if r["item"])
        fail_cnt = len(results_list) - ok_cnt

        m1, m2, m3 = st.columns(3)
        m1.metric("총 조회", f"{len(results_list)}개")
        m2.metric("✅ 성공", f"{ok_cnt}개")
        m3.metric("❌ 실패", f"{fail_cnt}개")

        if st.session_state.get("prdlst_excel_bytes"):
            fname = f"품목제조조회_{_dt2.now().strftime('%Y%m%d_%H%M')}.xlsx"
            dc1, dc2, dc3 = st.columns([1, 2, 1])
            with dc2:
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=st.session_state["prdlst_excel_bytes"],
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

        st.divider()

        for res in results_list:
            no   = res["no"]
            item = res.get("item")
            err  = res.get("error")

            if not item:
                st.error(f"❌ {no} — {err}")
                continue

            prod_nm = item.get("PRDLST_NM") or item.get("PRDLST_REPORT_NO") or no
            bssh_nm = item.get("BSSH_NM", "")
            svc_lbl = PRDLST_SERVICE_MAP.get(item.get("_svc_code", ""), "")

            with st.expander(f"🔎 {prod_nm}  [{svc_lbl}]", expanded=(ok_cnt == 1)):
                # 인허가 정보
                st.markdown("**인허가 정보**")
                ic1, ic2 = st.columns(2)
                ic1.markdown(f"**업체명** : {bssh_nm}")
                ic2.markdown(f"**소재지** : {item.get('SITE_ADDR','')}")
                st.divider()

                # 제품 정보
                st.markdown("**제품 정보**")
                r1c1, r1c2 = st.columns(2)
                r1c1.markdown(f"**품목보고번호** : {item.get('PRDLST_REPORT_NO','')}")
                r1c2.markdown(f"**일자** : {item.get('PRMS_DT','')}")
                r2c1, r2c2 = st.columns(2)
                r2c1.markdown(f"**제품명** : {item.get('PRDLST_NM','')}")
                r2c2.markdown(f"**소비기한** : {item.get('POG_DAYCNT','')}")
                r3c1, r3c2 = st.columns(2)
                r3c1.markdown(f"**유형** : {item.get('INDUTY_CD_NM','')}")
                r3c2.markdown(f"**식품유형(상세)** : {item.get('PRDLST_DCNM','')}")

                for label, key_nm in [("제품형태", "DISPOS"), ("포장재질", "PACKING_UNIT_DESCRIP"),
                                       ("내수/수출/겸용", "HIENG_LNTRT_DVS_NM"), ("용도", "USAGE"),
                                       ("생산여부", "PRODUCTION"), ("최종변경일", "CHNG_DT")]:
                    val = item.get(key_nm, "")
                    if val:
                        st.markdown(f"**{label}** : {val}")

                # 성분 및 원료
                ingredients = item.get("_ingredients", [])
                if ingredients:
                    st.divider()
                    st.markdown("**성분 및 원료**")
                    ing_data = [{"번호": i+1, "성분 및 원료": ing} for i, ing in enumerate(ingredients)]
                    st.dataframe(
                        pd.DataFrame(ing_data),
                        use_container_width=True,
                        hide_index=True,
                        height=min(40 * len(ingredients) + 40, 500),
                    )

# ============================================================
# main_tab4: 생산실적 조회
# ============================================================
PROD_API_CONFIG = {
    "식품/식품첨가물 생산실적": {"service_id": "I1320", "row_key": "I1320"},
    "축산물 생산실적":          {"service_id": "I1420", "row_key": "I1420"},
}

PROD_COLUMN_KR = {
    "BSSH_NM":             "업소명",
    "EVL_YR":              "보고년도",
    "LCNS_NO":             "인허가번호",
    "PRDLST_NM":           "품목명",
    "PRDLST_REPORT_NO":    "품목제조번호",
    "PRDCTN_QY":           "생산량(kg)",
    "PRDLST_CD_NM":        "품목유형",
    "GUBUN":               "구분",
    "H_ITEM_NM":           "대분류품목명",
    "FYER_PRDCTN_ABRT_QY": "연간생산중단수량",
    "PRMS_DT":             "허가일자",
    "PRMS_STTS":           "허가상태",
    "ENTP_NM":             "제조업체명",
    "PRDLST_CD":           "품목코드",
    "RAWMTRL_NM":          "원재료명",
    "PRDT_SHAP_CD_NM":     "제품형태",
    "LAST_UPDT_DTM":       "최종수정일",
    "CMPTN_PRDCTN_QY":     "완제품생산량(kg)",
    "WTSUPLY_PRDCTN_QY":   "위탁생산량(kg)",
    "SELF_PRDCTN_QY":      "자가생산량(kg)",
    "EXPRT_QY":            "수출량(kg)",
    "DOMST_SALE_QY":       "국내판매량(kg)",
    "MNFCTR_YY":           "제조연도",
    "INDUTY_NM":           "업종명",
    # 식품/식품첨가물 생산실적(I1320) 추가 필드
    "PRSDNT_NM":           "대표자명",
    "CLSBIZ_DVS_NM":       "폐업구분",
    "INSTT_NM":            "기관명",
    "LOCP_ADDR":           "소재지",
    "TELNO":               "전화번호",
    "SITE_ADDR":           "소재지",
    "INDUTY_CD_NM":        "업종유형",
    "PRDLST_DCNM":         "식품유형(상세)",
    "PRODUCTION":          "생산여부",
    "USAGE":               "용도",
    "POG_DAYCNT":          "소비기한",
    "DISPOS":              "제품형태(상세)",
    "PACKING_UNIT_DESCRIP":"포장재질",
    "HIENG_LNTRT_DVS_NM":  "내수/수출/겸용",
    "CHNG_DT":             "최종변경일",
}

# 통합 컬럼 출력 순서 (두 API 공통 기준)
PROD_COLUMN_ORDER = [
    "품목제조번호", "보고년도", "인허가번호", "품목명", "업소명",
    "생산량(kg)", "완제품생산량(kg)", "위탁생산량(kg)", "자가생산량(kg)",
    "수출량(kg)", "국내판매량(kg)", "품목유형", "대분류품목명", "구분",
    "연간생산중단수량", "허가일자", "허가상태", "제조업체명", "품목코드",
    "제품형태", "원재료명", "제조연도", "업종명", "최종수정일",
    # 식품/식품첨가물 전용
    "대표자명", "폐업구분", "기관명", "소재지", "전화번호",
    "업종유형", "식품유형(상세)", "생산여부", "용도", "소비기한",
    "제품형태(상세)", "포장재질", "내수/수출/겸용", "최종변경일",
]

def prod_fetch_all(api_key, service_id, row_key, range_start, range_end, extra_params):
    """생산실적 API 전체 수집 (1000건 단위 분할) - 클라이언트 필터링만 사용"""
    all_rows = []
    chunk = 1000
    cur = range_start
    while cur <= range_end:
        end_cur = min(cur + chunk - 1, range_end)
        # 식품안전나라 API는 URL 경로 파라미터 필터를 지원하지 않음 → 기본 URL만 사용
        url = (
            f"https://openapi.foodsafetykorea.go.kr/api"
            f"/{api_key}/{service_id}/json/{cur}/{end_cur}"
        )
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        rows = data.get(row_key, {}).get("row", [])
        all_rows.extend(rows)
        cur = end_cur + 1
    return all_rows

def build_prod_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="생산실적")
    return buf.getvalue()

with main_tab4:
    from datetime import datetime as _dt4

    st.markdown("### 📊 생산실적 조회 (식품안전나라)")
    st.info("API 키를 입력하고 조회 범위와 조건을 설정한 후 수집하세요. 최대 5,000건 이상 조회 가능합니다.")

    # ── API 키 ──
    with st.expander("🔑 API 인증키 설정", expanded=(not st.session_state.get("prod_api_key"))):
        prod_key_input = st.text_input(
            "식품안전나라 OpenAPI 인증키",
            value=st.session_state.get("prod_api_key", ""),
            type="password",
            placeholder="발급받은 인증키 입력",
            key="prod_api_key_input",
        )
        c_save, c_link = st.columns([1, 2])
        with c_save:
            if st.button("저장", key="prod_save_key"):
                st.session_state["prod_api_key"] = prod_key_input.strip()
                st.success("인증키 저장됨")
        with c_link:
            st.markdown("[식품안전나라 OpenAPI 신청 →](https://www.foodsafetykorea.go.kr/api/main.do)")

    # ── API 종류 선택 ──
    prod_api_name = st.selectbox("API 종류 선택", list(PROD_API_CONFIG.keys()), key="prod_api_name")
    prod_cfg = PROD_API_CONFIG[prod_api_name]

    # ── 범위 ──
    pc1, pc2 = st.columns(2)
    with pc1:
        prod_range_start = st.number_input("시작 번호", min_value=1, value=1, step=1, key="prod_range_start")
    with pc2:
        prod_range_end = st.number_input("끝 번호", min_value=1, value=1000, step=1, key="prod_range_end")

    # ── 선택 파라미터 ──
    with st.expander("🔧 파라미터 입력 (선택사항)"):
        pp1, pp2 = st.columns(2)
        with pp1:
            prod_evl_yr  = st.text_input("보고년도 (YYYY)", placeholder="예: 2023", key="prod_evl_yr")
            prod_prdlst  = st.text_input("품목명", placeholder="예: 양념육", key="prod_prdlst")
            prod_prdtype = st.text_input("품목유형", placeholder="예: 양념육(멸균)", key="prod_prdtype")
        with pp2:
            prod_bssh    = st.text_input("업소명", placeholder="예: 농심", key="prod_bssh")
            prod_lcns    = st.text_input("인허가번호", placeholder="숫자만 입력", key="prod_lcns")

    # ── 수집 버튼 ──
    if st.button("🔍 데이터 수집 시작", key="prod_fetch_btn", type="primary", use_container_width=True):
        _key = st.session_state.get("prod_api_key", "").strip()
        if not _key:
            st.warning("API 인증키를 먼저 저장하세요.")
        elif int(prod_range_start) > int(prod_range_end):
            st.warning("시작 번호가 끝 번호보다 큽니다.")
        else:
            _extra = {}
            if prod_evl_yr:  _extra["EVL_YR"]       = prod_evl_yr.strip()
            if prod_bssh:    _extra["BSSH_NM"]       = prod_bssh.strip()
            if prod_prdlst:  _extra["PRDLST_NM"]     = prod_prdlst.strip()
            if prod_lcns:    _extra["LCNS_NO"]        = prod_lcns.strip()
            if prod_prdtype: _extra["PRDLST_CD_NM"]  = prod_prdtype.strip()

            with st.spinner(f"{prod_api_name} 데이터 수집 중... ({int(prod_range_start)}~{int(prod_range_end)}번)"):
                try:
                    rows = prod_fetch_all(
                        _key,
                        prod_cfg["service_id"],
                        prod_cfg["row_key"],
                        int(prod_range_start),
                        int(prod_range_end),
                        _extra,
                    )
                    if rows:
                        df_prod = pd.DataFrame(rows)
                        df_prod.rename(columns={c: PROD_COLUMN_KR.get(c, c) for c in df_prod.columns}, inplace=True)
                        total_fetched = len(df_prod)
                        # ── 클라이언트 사이드 필터링 (API 서버는 URL 파라미터 필터 미지원) ──
                        if prod_prdlst and "품목명" in df_prod.columns:
                            df_prod = df_prod[df_prod["품목명"].str.contains(prod_prdlst.strip(), na=False)]
                        if prod_evl_yr and "보고년도" in df_prod.columns:
                            df_prod = df_prod[df_prod["보고년도"].astype(str).str.contains(prod_evl_yr.strip(), na=False)]
                        if prod_bssh and "업소명" in df_prod.columns:
                            df_prod = df_prod[df_prod["업소명"].str.contains(prod_bssh.strip(), na=False)]
                        df_prod = df_prod.reset_index(drop=True)
                        # 통합 컬럼 순서 적용
                        ordered_cols = [c for c in PROD_COLUMN_ORDER if c in df_prod.columns]
                        extra_cols   = [c for c in df_prod.columns if c not in PROD_COLUMN_ORDER]
                        df_prod = df_prod[ordered_cols + extra_cols]
                        st.session_state["prod_df"]    = df_prod
                        st.session_state["prod_excel"] = build_prod_excel(df_prod)
                        if len(df_prod) == 0:
                            filter_used = [k for k in [prod_prdlst, prod_evl_yr, prod_bssh] if k]
                            st.warning(
                                f"조회 범위 {int(prod_range_start)}~{int(prod_range_end)}에서 수집된 {total_fetched:,}건 중 "
                                f"조건에 맞는 항목이 없습니다.\n\n"
                                f"**해결 방법:** 끝 번호를 늘리거나(예: 5000), 품목명 철자를 확인하거나, "
                                f"API 종류(식품/식품첨가물 ↔ 축산물)를 바꿔보세요."
                            )
                        else:
                            total_msg = f"총 **{len(df_prod):,}건** 수집 완료!"
                            if len(df_prod) < total_fetched:
                                total_msg += f" (전체 {total_fetched:,}건 중 조건 필터 적용)"
                            st.success(total_msg)
                    else:
                        st.session_state["prod_df"]    = None
                        st.session_state["prod_excel"] = None
                        st.warning("조회 결과가 없습니다. API 키 또는 파라미터를 확인하세요.")
                except Exception as e:
                    st.error(f"오류 발생: {e}")

    # ── 결과 표시 ──
    if st.session_state.get("prod_df") is not None:
        df_show = st.session_state["prod_df"]

        pm1, pm2 = st.columns(2)
        pm1.metric("수집 건수", f"{len(df_show):,}건")
        pm2.metric("컬럼 수", f"{len(df_show.columns)}개")

        # 엑셀 다운로드
        fname_prod = f"생산실적_{prod_api_name.replace('/', '_').replace(' ', '')}_{_dt4.now().strftime('%Y%m%d_%H%M')}.xlsx"
        dc1, dc2, dc3 = st.columns([1, 2, 1])
        with dc2:
            st.download_button(
                label="📥 엑셀 다운로드",
                data=st.session_state["prod_excel"],
                file_name=fname_prod,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        st.divider()
        st.markdown("**미리보기 (상위 100행)**")
        st.dataframe(df_show.head(100), use_container_width=True, hide_index=True)


# ============================================================
# main_tab5: 월별예상금액 수집
# ============================================================
from scrapers.monthly_agent_scraper import MonthlyAgentScraper, build_combined_excel

# 세션 초기화
for _k, _v in {
    "monthly_summary_df": pd.DataFrame(),
    "monthly_company_results": [],
    "monthly_combined_excel": None,
    "monthly_done": False,
    "monthly_log": [],
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


async def _run_monthly(uid, pwd, sido, military, year, month, kw_inc, kw_exc, headless, log_list):
    async with MonthlyAgentScraper(uid, pwd, headless=headless) as scraper:
        def cb(msg):
            log_list.append(msg)
        summary_df, company_results = await scraper.collect(
            sido=sido, military=military, year=year, month=month,
            keyword_include=kw_inc, keyword_exclude=kw_exc,
            progress_callback=cb,
        )
    return summary_df, company_results


with main_tab5:
    from datetime import datetime as _dt5

    st.markdown("### 📋 월별예상금액 자동 수집 (FOODnBID)")
    st.info(
        "검색 조건을 설정하고 **수집 시작** 버튼을 누르면, 상위품목확인 업체별 Excel을 자동으로 "
        "다운로드해 하나의 파일로 합쳐 드립니다."
    )

    # ── 로그인 정보 ──
    with st.expander("🔐 FOODnBID 로그인 정보", expanded=not st.session_state.get("monthly_uid")):
        m_uid = st.text_input("아이디", key="monthly_uid_input", placeholder="FOODnBID 아이디")
        m_pwd = st.text_input("비밀번호", key="monthly_pwd_input", type="password", placeholder="비밀번호")
        if st.button("저장", key="monthly_save_cred"):
            st.session_state["monthly_uid"] = m_uid.strip()
            st.session_state["monthly_pwd"] = m_pwd.strip()
            st.success("로그인 정보 저장됨")

    st.divider()

    # ── 검색 조건 ──
    st.markdown("#### 🔍 검색 조건")
    mc1, mc2, mc3, mc4 = st.columns(4)

    with mc1:
        sido_options = {
            "전체": "",
            "서울": "11", "부산": "26", "대구": "27", "인천": "28",
            "광주": "29", "대전": "30", "울산": "31", "세종": "36",
            "경기": "41", "강원": "42", "충북": "43", "충남": "44",
            "전북": "45", "전남": "46", "경북": "47", "경남": "48", "제주": "50",
        }
        sido_label = st.selectbox("시도", list(sido_options.keys()), index=0, key="m_sido")
        sido_val = sido_options[sido_label]

    with mc2:
        military_options = {"군부대 제외": "2", "군부대 포함": "1", "전체": ""}
        military_label = st.selectbox("군부대", list(military_options.keys()), index=0, key="m_military")
        military_val = military_options[military_label]

    with mc3:
        current_year = _dt5.now().year
        year_list = [str(y) for y in range(current_year, current_year - 5, -1)]
        year_val = st.selectbox("연도", year_list, index=0, key="m_year")

    with mc4:
        month_val = st.selectbox(
            "월", [str(m) for m in range(1, 13)],
            index=_dt5.now().month - 1, key="m_month"
        )

    mc5, mc6 = st.columns(2)
    with mc5:
        kw_inc = st.text_input("검색포함 키워드", placeholder="예: 치킨", key="m_kw_inc")
    with mc6:
        kw_exc = st.text_input("검색제외 키워드", placeholder="예: 소스", key="m_kw_exc")

    m_headless = st.toggle("백그라운드 실행 (헤드리스)", value=True, key="m_headless",
                           help="OFF 시 브라우저 창이 보입니다 (디버깅용)")

    st.divider()

    # ── 수집 버튼 ──
    if st.button("🚀 수집 시작", type="primary", use_container_width=True, key="monthly_run"):
        _uid = st.session_state.get("monthly_uid", "").strip()
        _pwd = st.session_state.get("monthly_pwd", "").strip()
        if not _uid or not _pwd:
            st.warning("로그인 정보를 먼저 저장하세요.")
        else:
            st.session_state["monthly_done"] = False
            st.session_state["monthly_log"] = []
            st.session_state["monthly_summary_df"] = pd.DataFrame()
            st.session_state["monthly_company_results"] = []
            st.session_state["monthly_combined_excel"] = None

            progress_log: list = []
            _result_container: dict = {}

            def _monthly_thread():
                if sys.platform == "win32":
                    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    summary_df, company_results = loop.run_until_complete(
                        _run_monthly(
                            _uid, _pwd, sido_val, military_val,
                            year_val, month_val, kw_inc, kw_exc,
                            m_headless, progress_log,
                        )
                    )
                    _result_container["summary_df"] = summary_df
                    _result_container["company_results"] = company_results
                    _result_container["combined_excel"] = (
                        build_combined_excel(company_results) if company_results else None
                    )
                    _result_container["ok"] = True
                except Exception as e:
                    logger.error(f"월별예상금액 수집 오류: {e}")
                    import traceback
                    progress_log.append(f"❌ 오류: {e}")
                    progress_log.append(traceback.format_exc())
                    _result_container["ok"] = False
                finally:
                    loop.close()

            with st.spinner("수집 중... 브라우저 자동화가 실행됩니다. 잠시 기다려 주세요."):
                t = threading.Thread(target=_monthly_thread, daemon=True)
                t.start()
                t.join()

            # 스레드 완료 후 메인 스레드에서 session_state 업데이트
            st.session_state["monthly_log"] = list(progress_log)
            st.session_state["monthly_summary_df"] = _result_container.get("summary_df", pd.DataFrame())
            st.session_state["monthly_company_results"] = _result_container.get("company_results", [])
            st.session_state["monthly_combined_excel"] = _result_container.get("combined_excel")
            st.session_state["monthly_done"] = True

            st.rerun()

    # ── 진행 로그 ──
    if st.session_state.get("monthly_log"):
        with st.expander("📜 수집 로그", expanded=not st.session_state.get("monthly_done")):
            for line in st.session_state["monthly_log"]:
                st.text(line)

    # ── 결과 표시 ──
    if st.session_state.get("monthly_done"):
        summary_df = st.session_state["monthly_summary_df"]
        company_results = st.session_state["monthly_company_results"]

        st.success(f"✅ 수집 완료 — {len(company_results)}개 업체 데이터")

        # 메트릭
        rm1, rm2, rm3 = st.columns(3)
        rm1.metric("업체 수", f"{len(company_results)}개")
        total_rows = sum(len(r["df"]) for r in company_results)
        rm2.metric("총 상품 행 수", f"{total_rows:,}건")
        dl_ok = sum(1 for r in company_results if r.get("excel_bytes"))
        rm3.metric("Excel 다운로드 성공", f"{dl_ok}/{len(company_results)}")

        # 통합 Excel 다운로드
        if st.session_state.get("monthly_combined_excel"):
            fname = f"월별예상금액_{year_val}년{month_val}월_{kw_inc}_{_dt5.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="📥 전체 Excel 다운로드 (업체별 시트 + 전체 시트)",
                data=st.session_state["monthly_combined_excel"],
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="monthly_dl_combined",
            )

        st.divider()

        # 요약 테이블
        if not summary_df.empty:
            st.markdown("#### 📊 검색 결과 요약")
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
            st.divider()

        # 업체별 미리보기
        if company_results:
            st.markdown("#### 🏢 업체별 상위품목 미리보기")
            company_tabs = st.tabs([r["company"][:20] for r in company_results])
            for tab, result in zip(company_tabs, company_results):
                with tab:
                    df_c = result["df"]
                    if df_c.empty:
                        st.warning("데이터 없음")
                    else:
                        st.dataframe(df_c, use_container_width=True, hide_index=True)

                    # 업체 개별 Excel 다운로드
                    if result.get("excel_bytes"):
                        st.download_button(
                            label=f"📥 {result['company']} Excel 다운로드",
                            data=result["excel_bytes"],
                            file_name=f"{result['company']}_{year_val}년{month_val}월.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"monthly_dl_{result['company']}",
                        )
                    else:
                        st.caption("이 업체의 Excel 다운로드는 실패했습니다 (테이블 데이터는 위에 표시됨)")
