# -*- coding: utf-8 -*-
"""
학교급식 시장규모 추출 스킬 v1.0
FoodnBid(info.foodnbid.com) 월별예상금액 페이지에서
제품 키워드별 월별 낙찰금액을 자동 수집합니다.

사용법:
  py -X utf8 학교급식규모.py 새우만두
  py -X utf8 학교급식규모.py 새우만두 2025
  py -X utf8 학교급식규모.py 새우만두 2025 --excel
"""

import sys, os, subprocess, asyncio, json
from datetime import datetime

# ── 의존성 자동 설치 ──────────────────────────────────────
def ensure_packages():
    for mod, pkg in [("playwright", "playwright"), ("openpyxl", "openpyxl")]:
        try:
            __import__(mod)
        except ImportError:
            print(f"[설치 중] {pkg}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])
    # playwright 브라우저 설치
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            p.chromium.executable_path  # 이미 설치된 경우 통과
    except Exception:
        subprocess.check_call([sys.executable, "-m", "playwright", "install", "chromium", "--with-deps"])

ensure_packages()

from playwright.async_api import async_playwright

# ── 설정 ──────────────────────────────────────────────────
FOODNBID_ID = os.environ.get("FOODNBID_ID", "roffj2709")
FOODNBID_PW = os.environ.get("FOODNBID_PW", "b2bsolution1@")
FOODNBID_URL = "https://info.foodnbid.com"

BLUESIS_ID  = "씨제이프레시웨이서울"
BLUESIS_PW  = "1234"
BLUESIS_URL = "https://market.bluesis.com/web/pc/main.php"

IS_CLOUD = bool(os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RAILWAY_PROJECT_ID"))

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# 클라우드(Railway)에서는 영구 볼륨(/app/data)에 저장해 재배포 후에도 유지되도록 함
DATA_DIR   = "/app/data" if IS_CLOUD else SCRIPT_DIR
LOG_DIR    = os.path.join(DATA_DIR, "log")
OUTPUT_DIR = os.path.join(DATA_DIR, "output")
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ── FoodnBid 로그인 ───────────────────────────────────────
async def login(page):
    """FoodnBid에 로그인하고 세션을 반환합니다."""
    await page.goto(FOODNBID_URL + "/", wait_until="domcontentloaded", timeout=30000)
    await page.wait_for_timeout(1000)

    # login_cm_id/login_cm_pwd 필드는 loginopen div(display:none) 안에 있음
    # JS로 부모 요소를 visible 처리 후 fill
    await page.evaluate("""() => {
        let el = document.getElementById("login_cm_id");
        if (!el) return;
        while (el) { el.style.display = ""; el = el.parentElement; if (!el) break; }
    }""")

    await page.fill("#login_cm_id", FOODNBID_ID)
    await page.fill("#login_cm_pwd", FOODNBID_PW)

    # 팝업 자동 닫기
    page.on("dialog", lambda d: asyncio.ensure_future(d.accept()))

    # fn_login() 호출 (로그인 버튼 onclick)
    await page.evaluate("fn_login()")
    await page.wait_for_timeout(5000)

    logged_in = "로그아웃" in await page.content() or "CJ제일제당" in await page.content()
    if not logged_in:
        raise RuntimeError("FoodnBid 로그인 실패. 아이디/비밀번호를 확인하세요.")

    print(f"  [OK] FoodnBid 로그인 성공")
    return page


# ── 월별 검색 ─────────────────────────────────────────────
async def search_month(page, year: int, month: int, keyword: str, exclude: str = "") -> dict:
    """
    FoodnBid 월별예상금액 페이지에서 특정 연월·키워드로 검색 후
    총금액 및 업체별 금액을 반환합니다. (군부대제외 적용, exclude 검색제외 키워드 지원)

    Returns:
        {
            "year": 2025, "month": 3,
            "total": 108951186,
            "companies": [{"rank":1, "company":"대상", "amount":52725429}, ...]
        }
    """
    await page.goto(FOODNBID_URL + "/agent/sellingAgent01.do",
                    wait_until="domcontentloaded", timeout=15000)
    await page.wait_for_timeout(800)

    await page.select_option("#s_year", str(year))
    await page.select_option("#s_month", str(month))
    await page.select_option("#gbn_k", "001")  # 군부대제외 - 학교급식 매출만 집계
    await page.fill("#keyword", keyword)
    if exclude:
        await page.fill("#not_keyword", exclude)

    # 검색 버튼 클릭
    await page.click('button:has-text("검 색")')
    await page.wait_for_timeout(3000)

    # 총금액 추출 (back_green.bold 클래스 셀)
    total_str = await page.evaluate("""() => {
        let cells = Array.from(document.querySelectorAll(".back_green.bold"));
        let tot = cells.find(c => c.textContent.trim().match(/^[0-9,]+$/));
        if (tot) return tot.textContent.trim();

        // fallback: 총금액 행 탐색
        for (let row of document.querySelectorAll("tr")) {
            if (row.textContent.includes("총금액")) {
                for (let td of row.querySelectorAll("td")) {
                    if (td.textContent.trim().match(/^[0-9,]+$/))
                        return td.textContent.trim();
                }
            }
        }
        return "0";
    }""")
    total = int(total_str.replace(",", "")) if total_str else 0

    # 업체별 내역 추출
    companies = await page.evaluate("""() => {
        let rows = Array.from(document.querySelectorAll("#tBid tr, table tr"));
        let data = [];
        for (let row of rows) {
            let tds = row.querySelectorAll("td");
            if (tds.length >= 3) {
                let no  = tds[0].textContent.trim();
                let co  = tds[1].textContent.trim();
                let amt = tds[2].textContent.trim().replace(/,/g, "");
                if (no.match(/^[0-9]+$/) && amt.match(/^[0-9]+$/)) {
                    data.push({rank: parseInt(no), company: co, amount: parseInt(amt)});
                }
            }
        }
        return data;
    }""")

    return {"year": year, "month": month, "total": total, "companies": companies}


# ── 연간 12개월 수집 ──────────────────────────────────────
async def collect_annual(keyword: str, year: int, exclude: str = "") -> list:
    """1월~12월 전체 데이터를 수집합니다."""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page    = await context.new_page()

        await login(page)

        results = []
        for month in range(1, 13):
            try:
                data = await search_month(page, year, month, keyword, exclude)
                results.append(data)
                total_fmt = f"{data['total']:,}"
                top = data["companies"][:3]
                top_str = ", ".join(f"{c['company']}({c['amount']//10000}만)" for c in top)
                print(f"  {year}년 {month:2d}월: {total_fmt:>15}원  [{top_str}]")
            except Exception as e:
                print(f"  {year}년 {month:2d}월: 오류 - {e}")
                results.append({"year": year, "month": month, "total": 0, "companies": []})

        await browser.close()
    return results


# ── 전체 시장 제품 TOP5 수집 ─────────────────────────────
async def collect_top_products(keyword: str, year: int, results: list, exclude: str = "") -> list:
    """
    FoodnBid에서 제품별 연간 추정 매출을 계산, 전체 시장 TOP5 제품을 반환합니다.

    로직:
    1. 상위 10개 업체에 대해 상위 3개 비방학월의 모달 데이터 수집
    2. 같은 업체 내 동일 정규화 제품명 → 합산 (예: 데리야끼바베큐폭립 <증정> = 데리야끼바베큐폭립)
    3. 다른 제품명 → 별도 취급 (한올푸드 베이비백립 ≠ 크림치즈에퐁!립)
    4. 각 (업체, 제품) 쌍에 대해 샘플 비중 × 업체 연간 매출 = 제품 연간 추정
    5. 전체 (업체, 제품) 기준 연간 추정 매출 내림차순 TOP5 반환
    """
    VACATION = {1, 7, 12}

    # 업체별 연간 합계·방학제외 월 금액 목록
    comp_annual  = {}
    comp_monthly = {}
    for r in results:
        for co in r.get("companies", []):
            comp_annual[co["company"]] = comp_annual.get(co["company"], 0) + co["amount"]
            if r["month"] not in VACATION:
                comp_monthly.setdefault(co["company"], []).append(co["amount"])

    # 제품 수집 대상: 연간 매출 상위 10개 업체
    top10_companies = [c for c, _ in sorted(comp_annual.items(), key=lambda x: -x[1])[:10]]

    # 상위 3개 비방학월 선택
    non_vac = sorted(
        [r for r in results if r["month"] not in VACATION and r["total"] > 0],
        key=lambda x: -x["total"]
    )
    sample_months = [r["month"] for r in non_vac[:3]]
    if not sample_months:
        return []

    # 샘플 집계: (업체, 정규화제품명) → 샘플 합산 금액
    sample_prod   = {}   # (company, norm_name) → sample_amount
    sample_comp   = {}   # company → 모달에서 집계된 총 샘플 금액

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page    = await context.new_page()
        await login(page)

        for month in sample_months:
            print(f"  [{month}월] 제품 수집 중...")
            await page.goto(FOODNBID_URL + "/agent/sellingAgent01.do",
                            wait_until="domcontentloaded", timeout=15000)
            await page.wait_for_timeout(800)
            await page.select_option("#s_year", str(year))
            await page.select_option("#s_month", str(month))
            await page.select_option("#gbn_k", "001")  # 군부대제외 - 학교급식 매출만 집계
            await page.fill("#keyword", keyword)
            if exclude:
                await page.fill("#not_keyword", exclude)
            await page.click('button:has-text("검 색")')
            await page.wait_for_timeout(3000)

            company_codes = await page.evaluate("""() => {
                let result = {};
                for (let row of document.querySelectorAll("tr")) {
                    let links = Array.from(row.querySelectorAll("a"));
                    let link = links.find(a =>
                        (a.getAttribute('onclick') || '').includes('fn_getBest2') ||
                        (a.getAttribute('href')    || '').includes('fn_getBest2')
                    );
                    if (!link) continue;
                    let attr = link.getAttribute('onclick') || link.getAttribute('href') || '';
                    let m = attr.match(/fn_getBest2\((\d+)/);
                    if (!m) continue;
                    let tds = row.querySelectorAll("td");
                    if (tds.length > 1) result[tds[1].textContent.trim()] = m[1];
                }
                return result;
            }""")

            for company in top10_companies:
                code = company_codes.get(company)
                if not code:
                    continue
                try:
                    await page.evaluate(f"fn_getBest2({code}, 1)")
                    await page.wait_for_timeout(1500)

                    modal_rows = await page.evaluate("""() => {
                        let modal = document.querySelector(".modal-content");
                        if (!modal) return [];
                        return Array.from(modal.querySelectorAll("tr"))
                            .map(r => Array.from(r.querySelectorAll("td"))
                                          .map(td => td.textContent.trim()))
                            .filter(cols => cols.length >= 3 && /^\\d+$/.test(cols[0]));
                    }""")

                    for cols in modal_rows:
                        prod_name  = cols[2] if len(cols) > 2 else ""
                        amount_str = cols[3] if len(cols) > 3 else "0"
                        try:
                            amount = int(amount_str.replace(",", ""))
                        except Exception:
                            amount = 0
                        if prod_name and amount > 0:
                            norm = normalize_product_name(prod_name)
                            if norm:
                                key = (company, norm)
                                sample_prod[key] = sample_prod.get(key, 0) + amount
                                sample_comp[company] = sample_comp.get(company, 0) + amount

                    await page.evaluate("""() => {
                        let btn = document.querySelector(
                            ".modal .close, .modal-header .close, [data-dismiss='modal']");
                        if (btn) btn.click();
                    }""")
                    await page.wait_for_timeout(400)

                except Exception as e:
                    print(f"    [{company}] {month}월 오류: {e}")

        await browser.close()

    # 각 (업체, 제품) 의 연간 추정 매출 계산
    # = 업체 연간 합계 × (해당 제품 샘플 합 / 업체 전체 샘플 합)
    prod_annual_est = {}
    for (company, norm_name), samp_amt in sample_prod.items():
        comp_samp = sample_comp.get(company, 0)
        annual_est = int(comp_annual.get(company, 0) * (samp_amt / comp_samp)) \
                     if comp_samp > 0 else 0
        prod_annual_est[(company, norm_name)] = annual_est

    # 전체 (업체, 제품) 기준 내림차순 정렬 → TOP5
    top5_pairs = sorted(prod_annual_est.items(), key=lambda x: -x[1])[:5]

    product_info = []
    for rank, ((company, norm_name), annual_est) in enumerate(top5_pairs, 1):
        months_list = comp_monthly.get(company, [])
        comp_samp   = sample_comp.get(company, 0)
        prod_samp   = sample_prod.get((company, norm_name), 0)
        prod_share  = prod_samp / comp_samp if comp_samp > 0 else 0
        monthly_avg = int((sum(months_list) / len(months_list)) * prod_share) \
                      if months_list else 0

        weight = _extract_weight(norm_name)
        origin = _extract_origin(norm_name)

        product_info.append({
            "rank":        rank,
            "company":     company,
            "product":     norm_name,
            "products":    [{"name": norm_name, "amount": annual_est}],
            "annual":      annual_est,
            "monthly_avg": monthly_avg,
            "weight":      weight,
            "origin":      origin,
        })
        print(f"  [{rank}위] {company} - {norm_name}: 연간추정 {annual_est:,}원")

    return product_info


# ── 블루시스 가격 수집 ────────────────────────────────────
_BLUESIS_JS_ROWS = """
() => {
    var brandEls  = Array.from(document.querySelectorAll(".w_brand"));
    var res = [];
    for (var i = 0; i < brandEls.length; i++) {
        var row = brandEls[i].parentElement;
        if (!row) continue;
        var comEl    = row.querySelector(".w_com");
        var pnameEl  = row.querySelector(".w_pname");
        var kpriceEl = row.querySelector(".w_kprice");
        var b = brandEls[i].innerText.split("\\n")[0].replace("가입","").trim();
        var c = comEl    ? comEl.innerText.split("\\n")[0].replace("가입","").trim() : "";
        var price = "";
        if (kpriceEl) {
            var t = kpriceEl.innerText.replace("학교 kg단가","").trim();
            var m = t.match(/([0-9]{1,3}(?:,[0-9]{3})+)/);
            price = m ? m[1] + "원/kg" : "싯가";
        }
        var pn = "";
        if (pnameEl) {
            var clone = pnameEl.cloneNode(true);
            var btn = clone.querySelector("button");
            if (btn) btn.remove();
            pn = clone.innerText.trim().split("\\n")[0].slice(0, 80);
        }
        res.push({brand: b, com: c, pname: pn, kprice: price});
    }
    return res;
}
"""


async def collect_bluesis_prices(keyword: str, product_info: list) -> dict:
    """
    블루시스(market.bluesis.com)에 로그인하여 키워드 검색 후
    경쟁사 TOP5의 학교kg단가를 수집합니다.

    - 로그인: login.php → #blue_uid / #pwd / input[value='로그인하기']
    - 검색: product.php?from=main&_qr={keyword}
    - 건수: #rows select → 100건
    - 파싱: .w_brand 부모 컨테이너 기준, .w_kprice에서 가격 추출
    - 매칭: product_info의 origin(국내산/수입산)과 일치하는 제품 우선 선택

    Returns:
        {company: "XX,XXX원/kg"} 형태의 딕셔너리
    """
    import urllib.parse
    company_names = [p["company"] for p in product_info]
    origin_map    = {p["company"]: p.get("origin", "") for p in product_info}
    prices = {name: "직접 확인 필요" for name in company_names}

    LOGIN_URL   = "https://market.bluesis.com/web/pc/login.php"
    SEARCH_URL  = (
        "https://market.bluesis.com/web/pc/product.php"
        f"?from=main&_qr={urllib.parse.quote(keyword)}"
    )

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        page    = await browser.new_page()

        try:
            # ── 로그인 ──
            print(f"  [블루시스] 로그인 중...")
            await page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
            await page.wait_for_timeout(1500)
            await page.fill("#blue_uid", BLUESIS_ID)
            await page.fill("#pwd",      BLUESIS_PW)
            await page.click("input[value='로그인하기']")
            await page.wait_for_timeout(3000)
            logged_in = "로그아웃" in await page.evaluate("document.body.innerText")
            print(f"  [블루시스] {'로그인 성공' if logged_in else '로그인 상태 불명확'}")

            # ── 검색 페이지 이동 + 100건 로드 ──
            await page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=15000)
            await page.wait_for_timeout(2000)
            try:
                await page.select_option("#rows", "100")
                await page.wait_for_timeout(3000)
            except Exception:
                pass
            print(f"  [블루시스] '{keyword}' 검색 완료")

            # ── 전체 제품 파싱 ──
            items = await page.evaluate(_BLUESIS_JS_ROWS)
            print(f"  [블루시스] {len(items)}건 파싱")

            # ── 업체별 매칭: origin 일치 제품 우선 ──
            for company in company_names:
                want_origin = origin_map.get(company, "")
                matches = [
                    it for it in items
                    if company in it["brand"] or company in it["com"]
                ]
                if not matches:
                    prices[company] = "블루시스 미등록"
                    print(f"  [블루시스] {company}: 미등록")
                    continue

                # origin 일치 + 실제 가격 있는 것 우선
                # origin이 명시적으로 반대인 제품(수입산을 원하는데 국내산이 pname에 있거나 vice versa)은 하위 순위
                opposite = {"국내산": "수입산", "수입산": "국내산"}.get(want_origin, "")
                def score(it):
                    pn        = it["pname"]
                    has_price = it["kprice"] not in ("", "싯가")
                    origin_hit  = bool(want_origin and want_origin in pn)
                    origin_miss = bool(opposite and opposite in pn)
                    return (origin_hit and has_price,
                            has_price and not origin_miss,
                            origin_hit,
                            not origin_miss)

                best = max(matches, key=score)
                prices[company] = best["kprice"] if best["kprice"] not in ("", "싯가") else "싯가"
                print(f"  [블루시스] {company}: {prices[company]}  ({best['pname'][:40]})")

        except Exception as e:
            print(f"  [블루시스] 수집 오류: {e}")
        finally:
            await browser.close()

    return prices


def _extract_weight(product_name: str) -> str:
    import re
    m = re.search(r'(\d+(?:\.\d+)?)\s*(kg|g|ml|L)', product_name, re.IGNORECASE)
    return f"{m.group(1)}{m.group(2)}" if m else "확인 필요"


def _extract_origin(product_name: str) -> str:
    if any(k in product_name for k in ("국내산", "국산", "우리돼지", "우리닭", "국내")):
        return "국내산"
    if any(k in product_name for k in ("수입산", "수입", "외국산")):
        return "수입산"
    return "확인 필요"


def normalize_product_name(name: str) -> str:
    """제품명에서 용량·규격·프로모션 텍스트를 제거하고 핵심 이름만 반환."""
    import re
    name = re.sub(r'\s*//.*', '', name)                                          # // 이후 제거
    name = re.sub(r'\*\*[^*]*\*\*', '', name)                                   # **강조** 제거
    name = re.sub(r'\s*<[^>]*>', '', name)                                       # <프로모션> 제거
    name = re.sub(r'\s*[\(\[（【][^\)\]）】]*[\)\]）】]', '', name)              # 괄호 내용 제거
    name = re.sub(r'\s+\d+(?:\.\d+)?(?:G|g|KG|Kg|kg|ML|ml|L|l)\b', '', name)  # 중량 제거
    name = re.sub(r'\s+', ' ', name).strip()
    return name


# ── 결과 출력 ─────────────────────────────────────────────
MONTH_NAMES = ["1월","2월","3월","4월","5월","6월",
               "7월","8월","9월","10월","11월","12월"]
VACATION    = {1, 7, 12}   # 방학월

def print_summary(keyword: str, year: int, results: list):
    annual = sum(r["total"] for r in results)
    best_month = max(results, key=lambda x: x["total"])["month"]
    print()
    print("=" * 65)
    print(f"  {keyword}  FoodnBid 수도권 월별 낙찰금액  ({year}년)")
    print("=" * 65)
    for r in results:
        flag = " ★" if r["month"] == best_month else ""
        bar_len = int(r["total"] / max(1, annual) * 40)
        bar = "█" * bar_len
        print(f"  {MONTH_NAMES[r['month']-1]:>4}: {r['total']:>13,}원  {bar}{flag}")
    print("-" * 65)
    print(f"  연간 합계: {annual:>13,}원  ({annual/1e8:.2f}억원)")
    print()

    # 시장규모 환산
    coverage     = 0.60
    school_suw   = int(annual / coverage)
    school_nation = int(school_suw / 0.405)
    dangtche     = int(school_nation * 0.225)
    total_c      = school_nation + dangtche

    print("  ── 학교급식 시장 규모 환산 ──")
    print(f"  ① 수도권 FoodnBid 연간 실측 : {annual:>13,}원")
    print(f"  ② ÷ FoodnBid 커버리지 60%  → 수도권 학교급식 : {school_suw:>13,}원")
    print(f"  ③ ÷ 수도권 학생비율 40.5%  → 전국 학교급식   : {school_nation:>13,}원  ({school_nation/1e8:.1f}억)")
    print(f"  ④ + 단체급식 22.5%          → 급식 전체 합계  : {total_c:>13,}원  ({total_c/1e8:.0f}억)")
    print("=" * 65)


# ── 엑셀 저장 (3개 시트) ──────────────────────────────────
def save_excel(keyword: str, year: int, results: list, prev_results=None):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [주의] openpyxl 없음 - 엑셀 저장 생략")
        return

    # ── 스타일 상수 ────────────────────────────────────────
    thin   = Side(style="thin",   color="BFBFBF")
    medium = Side(style="medium", color="2E5F8E")
    BD  = Border(left=thin, right=thin, top=thin, bottom=thin)
    BDM = Border(left=medium, right=medium, top=medium, bottom=medium)

    F_BLUE   = PatternFill("solid", fgColor="2E5F8E")  # 진파랑 - 제목/헤더
    F_LBLUE  = PatternFill("solid", fgColor="D6E4F0")  # 연파랑 - 레이블
    F_GREEN  = PatternFill("solid", fgColor="EBF5E1")  # 연초록 - 합계
    F_VAC    = PatternFill("solid", fgColor="E8F0FE")  # 하늘   - 방학월
    F_ORANGE = PatternFill("solid", fgColor="FFF2CC")  # 연노랑 - 강조
    F_GRAY   = PatternFill("solid", fgColor="F2F2F2")  # 연회색 - 참고
    F_KPINK  = PatternFill("solid", fgColor="FCE4EC")  # 연분홍 - 1위
    F_WHITE  = PatternFill()

    FT_TITLE = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
    FT_HEAD  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    FT_LABEL = Font(name="맑은 고딕", bold=True, color="1F4E79", size=10)
    FT_SUM   = Font(name="맑은 고딕", bold=True, color="375623", size=10)
    FT_BOLD      = Font(name="맑은 고딕", bold=True, size=10)
    FT_N         = Font(name="맑은 고딕", size=10)
    FT_HEAD_PINK = Font(name="맑은 고딕", bold=True, color="1F4E79", size=10)  # 분홍배경 1위 헤더용
    FT_S     = Font(name="맑은 고딕", size=9)
    FT_SRC   = Font(name="맑은 고딕", size=8, italic=True, color="666666")

    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AR = Alignment(horizontal="right",  vertical="center")

    # 사전 계산
    prev_results   = prev_results or []
    prev_map       = {r["month"]: r["total"] for r in prev_results}
    prev_annual    = sum(r["total"] for r in prev_results)
    prev_year      = year - 1

    annual         = sum(r["total"] for r in results)
    coverage       = 0.60
    school_suw     = int(annual / coverage)
    school_nation  = int(school_suw / 0.405)
    dangtche       = int(school_nation * 0.225)
    total_c        = school_nation + dangtche

    prev_school_suw    = int(prev_annual / coverage) if prev_annual else 0
    prev_school_nation = int(prev_school_suw / 0.405) if prev_annual else 0
    if prev_school_nation > 0:
        yoy_pct_nat = round((school_nation - prev_school_nation) / prev_school_nation * 100, 1)
        yoy_nat_str = f"{'▲' if yoy_pct_nat >= 0 else '▼'} {abs(yoy_pct_nat):.1f}%"
    else:
        yoy_nat_str = "N/A"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 기본 시트 제거

    # ════════════════════════════════════════════════════════
    #  SHEET 1 : 학교급식_월별실측
    # ════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("학교급식_월별실측")
    ws1.sheet_view.showGridLines = False
    # B=월 C=현년실측 D=비중 E=전년실측 F=전년동월대비 G=TOP3 H=현년전국환산(차트) I=전년전국환산(차트)
    for col, w in {1:2, 2:10, 3:20, 4:10, 5:20, 6:14, 7:42, 8:18, 9:18}.items():
        ws1.column_dimensions[get_column_letter(col)].width = w

    # 제목
    ws1.row_dimensions[1].height = 38
    ws1.merge_cells("B1:I1")
    c = ws1.cell(1, 2, f"【{keyword}】  학교급식 FoodnBid 월별 낙찰금액  ({year}년 vs {prev_year}년)")
    c.fill = F_BLUE; c.font = FT_TITLE; c.alignment = AC; c.border = BD

    # 요약 박스 (6 KPI: B~G)
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 26
    sum_labels = ["수도권 FoodnBid 연간", "수도권 학교급식 추정", "전국 학교급식 추정",
                  "전국 급식 전체", f"{prev_year}년 전국급식", "전년 대비 성장"]
    sum_vals   = [f"{annual/1e8:.2f}억원", f"{school_suw/1e8:.2f}억원",
                  f"{school_nation/1e8:.1f}억원", f"약 {total_c/1e8:.0f}억원",
                  f"{prev_school_nation/1e8:.1f}억원" if prev_school_nation else "N/A",
                  yoy_nat_str]
    sum_fills  = [F_LBLUE, F_LBLUE, F_GREEN, F_ORANGE, F_GRAY, F_ORANGE]
    sum_fonts  = [FT_LABEL, FT_LABEL, FT_SUM, FT_BOLD, FT_N, FT_BOLD]
    for j, (lbl, val, fill, fnt) in enumerate(zip(sum_labels, sum_vals, sum_fills, sum_fonts), 2):
        cl = ws1.cell(2, j, lbl)
        cl.fill = fill; cl.font = FT_S; cl.alignment = AC; cl.border = BD
        cv = ws1.cell(3, j, val)
        cv.fill = fill; cv.font = fnt; cv.alignment = AC; cv.border = BD

    # 헤더 행 (H, I는 차트 데이터용 - 연회색)
    ws1.row_dimensions[4].height = 14
    ws1.row_dimensions[5].height = 36
    main_headers = [
        "월",
        f"FoodnBid 실측\n({year}, 수도권, 원)",
        "비중\n(%)",
        f"FoodnBid 실측\n({prev_year}, 수도권, 원)",
        "전년\n동월 대비",
        "경쟁사 TOP 3 (낙찰금액)",
    ]
    for j, h in enumerate(main_headers, 2):
        c = ws1.cell(5, j, h)
        c.fill = F_BLUE; c.font = FT_HEAD; c.alignment = AC; c.border = BD
    for j, h in [(8, f"{year} 전국학교급식 환산"), (9, f"{prev_year} 전국학교급식 환산")]:
        c = ws1.cell(5, j, h)
        c.fill = F_GRAY; c.font = FT_S; c.alignment = AC; c.border = BD

    # 월별 데이터
    max_val    = max(x["total"] for x in results)
    best_month = next(r["month"] for r in results if r["total"] == max_val)
    for r in results:
        row = r["month"] + 5
        ws1.row_dimensions[row].height = 20
        is_vac   = r["month"] in VACATION
        is_best  = r["month"] == best_month
        row_fill = F_VAC if is_vac else F_WHITE
        pct      = round(r["total"] / annual * 100, 1) if annual else 0
        prev_m   = prev_map.get(r["month"], 0)

        # 전년 동월 대비
        if prev_m > 0:
            yoy_m    = round((r["total"] - prev_m) / prev_m * 100, 1)
            yoy_str  = f"{'▲' if yoy_m >= 0 else '▼'} {abs(yoy_m):.1f}%"
            yoy_fill = PatternFill("solid", fgColor="E8FFE8") if yoy_m >= 0 else PatternFill("solid", fgColor="FFE8E8")
        else:
            yoy_str, yoy_fill = "N/A", F_WHITE

        # 월
        c2 = ws1.cell(row, 2, MONTH_NAMES[r["month"]-1] + (" ★" if is_best else ""))
        c2.fill = row_fill; c2.font = FT_BOLD if is_best else FT_N; c2.alignment = AC; c2.border = BD

        # 현년 실측
        c3 = ws1.cell(row, 3, r["total"])
        c3.fill = row_fill; c3.font = FT_BOLD if r["total"] == max_val else FT_N
        c3.alignment = AR; c3.border = BD; c3.number_format = "#,##0"

        # 비중
        c4 = ws1.cell(row, 4, pct / 100)
        c4.fill = row_fill; c4.font = FT_N; c4.alignment = AC
        c4.border = BD; c4.number_format = "0.0%"

        # 전년 실측
        c5 = ws1.cell(row, 5, prev_m if prev_m else None)
        c5.fill = row_fill; c5.font = FT_N; c5.alignment = AR
        c5.border = BD; c5.number_format = "#,##0"

        # 전년 동월 대비
        c6 = ws1.cell(row, 6, yoy_str)
        c6.fill = yoy_fill; c6.font = FT_S; c6.alignment = AC; c6.border = BD

        # TOP 3
        top = r["companies"][:3]
        top_str = " | ".join(
            f"{'①' if i==0 else '②' if i==1 else '③'} {co['company']} {co['amount']:,}원"
            for i, co in enumerate(top)
        ) if top else "낙찰 없음"
        c7 = ws1.cell(row, 7, top_str)
        c7.fill = row_fill; c7.font = FT_S; c7.alignment = AL; c7.border = BD

        # 차트용: 전국 학교급식 환산값 (H, I)
        curr_nation_m = int(r["total"] / coverage / 0.405)
        prev_nation_m = int(prev_m / coverage / 0.405) if prev_m else None
        c8 = ws1.cell(row, 8, curr_nation_m)
        c8.fill = F_GRAY; c8.font = FT_S; c8.alignment = AR
        c8.border = BD; c8.number_format = "#,##0"
        c9 = ws1.cell(row, 9, prev_nation_m)
        c9.fill = F_GRAY; c9.font = FT_S; c9.alignment = AR
        c9.border = BD; c9.number_format = "#,##0"

    # 연간 합계 행
    r_sum = 18
    ws1.row_dimensions[r_sum].height = 24
    c_sl = ws1.cell(r_sum, 2, "연간 합계")
    c_sl.fill = F_GREEN; c_sl.font = FT_SUM; c_sl.alignment = AC; c_sl.border = BD

    c3s = ws1.cell(r_sum, 3, annual)
    c3s.fill = F_GREEN; c3s.font = FT_SUM; c3s.alignment = AR; c3s.border = BD; c3s.number_format = "#,##0"

    c4s = ws1.cell(r_sum, 4, 1.0)
    c4s.fill = F_GREEN; c4s.font = FT_SUM; c4s.alignment = AC; c4s.border = BD; c4s.number_format = "0.0%"

    c5s = ws1.cell(r_sum, 5, prev_annual if prev_annual else None)
    c5s.fill = F_GREEN; c5s.font = FT_SUM; c5s.alignment = AR; c5s.border = BD; c5s.number_format = "#,##0"

    ws1.merge_cells(f"F{r_sum}:G{r_sum}")
    ces = ws1.cell(r_sum, 6, f"연간 {annual/1e8:.2f}억원  |  전년({prev_year}) {prev_annual/1e8:.2f}억원  (수도권 FoodnBid 기준)")
    ces.fill = F_GREEN; ces.font = FT_SUM; ces.alignment = AL; ces.border = BD

    c8s = ws1.cell(r_sum, 8, school_nation)
    c8s.fill = F_GREEN; c8s.font = FT_SUM; c8s.alignment = AR; c8s.border = BD; c8s.number_format = "#,##0"
    c9s = ws1.cell(r_sum, 9, prev_school_nation if prev_school_nation else None)
    c9s.fill = F_GREEN; c9s.font = FT_SUM; c9s.alignment = AR; c9s.border = BD; c9s.number_format = "#,##0"

    # 출처 행
    ws1.row_dimensions[19].height = 14
    ws1.merge_cells("B19:I19")
    src = ws1.cell(19, 2, f"※ 출처: FoodnBid info.foodnbid.com 월별예상금액 (수도권·서울·경기·인천 학교) | 수집일: {datetime.now().strftime('%Y-%m-%d')}")
    src.fill = F_GRAY; src.font = FT_SRC; src.alignment = AL; src.border = BD

    # ── 꺾은선 차트 (전국 학교급식 환산값, 2개 연도 비교) ──────
    chart = LineChart()
    chart.title    = f"{keyword}  월별 전국 학교급식 시장규모 환산  ({year} vs {prev_year})"
    chart.y_axis.title = "전국 학교급식 환산 (원)"
    chart.x_axis.title = "월"
    chart.style    = 10
    chart.width    = 26
    chart.height   = 14
    chart.grouping = "standard"

    ref_curr = Reference(ws1, min_col=8, min_row=5, max_row=17)
    ref_prev = Reference(ws1, min_col=9, min_row=5, max_row=17)
    cats_ref = Reference(ws1, min_col=2, min_row=6, max_row=17)
    chart.add_data(ref_curr, titles_from_data=True)
    chart.add_data(ref_prev, titles_from_data=True)
    chart.set_categories(cats_ref)

    # 2025 — 진파랑, 두꺼운 선, 원형 마커
    s0 = chart.series[0]
    s0.graphicalProperties.line.solidFill  = "1F4E79"
    s0.graphicalProperties.line.width      = 28800        # 2.25pt
    s0.marker.symbol   = "circle"
    s0.marker.size     = 7
    s0.marker.graphicalProperties.fgColor  = "1F4E79"
    s0.marker.graphicalProperties.solidFill = "1F4E79"
    s0.smooth = False

    # 2024 — 주황, 두꺼운 선, 다이아몬드 마커
    s1 = chart.series[1]
    s1.graphicalProperties.line.solidFill  = "E8A838"
    s1.graphicalProperties.line.width      = 28800
    s1.marker.symbol   = "diamond"
    s1.marker.size     = 7
    s1.marker.graphicalProperties.fgColor  = "E8A838"
    s1.marker.graphicalProperties.solidFill = "E8A838"
    s1.smooth = False

    # 제목 폰트 크기 (16pt)
    try:
        from openpyxl.drawing.text import CharacterProperties
        chart.tx.rich.p[0].pPr.defRPr = CharacterProperties(sz=1600, b=True)
    except Exception:
        pass

    ws1.add_chart(chart, "B21")

    # ════════════════════════════════════════════════════════
    #  SHEET 2 : 시장규모_환산
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("시장규모_환산")
    ws2.sheet_view.showGridLines = False
    # B=항목 C=현년수치 D=현년억원 E=전년수치 F=전년억원 G=근거
    for col, w in {1:2, 2:30, 3:16, 4:12, 5:16, 6:12, 7:32}.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    # 전년 추가 계산
    prev_dangtche = int(prev_school_nation * 0.225) if prev_school_nation else 0
    prev_total_c  = prev_school_nation + prev_dangtche

    # YoY 헬퍼
    def _yoy_s(curr, prev):
        if not prev:
            return "N/A"
        p = round((curr - prev) / prev * 100, 1)
        return f"{'▲' if p >= 0 else '▼'} {abs(p):.1f}%"

    def _yoy_f(curr, prev):
        if not prev:
            return F_WHITE
        return PatternFill("solid", fgColor="E8FFE8") if curr >= prev else PatternFill("solid", fgColor="FFE8E8")

    # 제목
    ws2.row_dimensions[1].height = 38
    ws2.merge_cells("B1:G1")
    c = ws2.cell(1, 2, f"【{keyword}】  학교급식 시장 규모 환산  ({year}년 vs {prev_year}년 비교)")
    c.fill = F_BLUE; c.font = FT_TITLE; c.alignment = AC; c.border = BD

    # 요약 박스 헤더 (row 3)
    ws2.row_dimensions[2].height = 10
    ws2.row_dimensions[3].height = 30
    sum_hdr = ["구분", f"{year}년 수치", f"{year}년", f"{prev_year}년 수치", f"{prev_year}년", "비고"]
    for j, h in enumerate(sum_hdr, 2):
        c = ws2.cell(3, j, h)
        c.fill = F_BLUE; c.font = FT_HEAD; c.alignment = AC; c.border = BD

    # 요약 KPI rows (4~7)
    sum_rows = [
        ("전국 급식 전체 (최종 추정)", total_c,        f"약 {total_c/1e8:.0f}억원",
         prev_total_c,   f"약 {prev_total_c/1e8:.0f}억원" if prev_total_c else "N/A",
         F_ORANGE, FT_BOLD),
        ("전국 학교급식",              school_nation,   f"{school_nation/1e8:.1f}억원",
         prev_school_nation, f"{prev_school_nation/1e8:.1f}억원" if prev_school_nation else "N/A",
         F_GREEN, FT_SUM),
        ("수도권 학교급식",            school_suw,      f"{school_suw/1e8:.2f}억원",
         prev_school_suw, f"{prev_school_suw/1e8:.2f}억원" if prev_school_suw else "N/A",
         F_LBLUE, FT_LABEL),
        ("FoodnBid 실측 (수도권)",     annual,          f"{annual/1e8:.2f}억원",
         prev_annual, f"{prev_annual/1e8:.2f}억원" if prev_annual else "N/A",
         F_LBLUE, FT_LABEL),
    ]
    for i, (lbl, cv, cs, pv, ps, fill, fnt) in enumerate(sum_rows, 4):
        ws2.row_dimensions[i].height = 28
        cl = ws2.cell(i, 2, lbl)
        cl.fill = fill; cl.font = fnt; cl.alignment = AL; cl.border = BD
        cc = ws2.cell(i, 3, cv)
        cc.fill = fill; cc.font = fnt; cc.alignment = AR; cc.border = BD; cc.number_format = "#,##0"
        ccs = ws2.cell(i, 4, cs)
        ccs.fill = fill; ccs.font = fnt; ccs.alignment = AC; ccs.border = BD
        cp = ws2.cell(i, 5, pv if pv else None)
        cp.fill = F_GRAY; cp.font = FT_N; cp.alignment = AR; cp.border = BD; cp.number_format = "#,##0"
        cps = ws2.cell(i, 6, ps)
        cps.fill = F_GRAY; cps.font = FT_N; cps.alignment = AC; cps.border = BD
        ws2.cell(i, 7).border = BD

    # 단계별 환산 계산표 (row 10 헤더, 11~15 데이터)
    ws2.row_dimensions[9].height = 10
    ws2.row_dimensions[10].height = 32
    calc_headers = ["계산 단계", f"{year}년 수치 (원)", f"{year}년 억원",
                    f"{prev_year}년 수치 (원)", f"{prev_year}년 억원", "산출 근거·가정"]
    for j, h in enumerate(calc_headers, 2):
        c = ws2.cell(10, j, h)
        c.fill = F_BLUE; c.font = FT_HEAD; c.alignment = AC; c.border = BD

    calc_rows = [
        ("① 수도권 FoodnBid 연간 실측",
         annual,        f"{annual/1e8:.2f}억원",
         prev_annual,   f"{prev_annual/1e8:.2f}억원" if prev_annual else "N/A",
         f"FoodnBid info.foodnbid.com 실측 ({year}년 / {prev_year}년 1~12월)"),
        ("② ÷ FoodnBid 커버리지 60%  →  수도권 학교급식 시장",
         school_suw,    f"{school_suw/1e8:.2f}억원",
         prev_school_suw, f"{prev_school_suw/1e8:.2f}억원" if prev_school_suw else "N/A",
         "FoodnBid가 수도권 학교급식의 약 60% 커버 (직계약·수의계약 제외)"),
        ("③ ÷ 수도권 학생비율 40.5%  →  전국 학교급식 시장",
         school_nation, f"{school_nation/1e8:.1f}억원",
         prev_school_nation, f"{prev_school_nation/1e8:.1f}억원" if prev_school_nation else "N/A",
         "서울+경기 학생 수 / 전국 (교육통계서비스 2024)"),
        ("④ 전국 단체급식 추가 (× 22.5%)",
         dangtche,      f"{dangtche/1e8:.1f}억원",
         prev_dangtche, f"{prev_dangtche/1e8:.1f}억원" if prev_dangtche else "N/A",
         "군·병원·기업 단체급식, 학교급식 대비 약 22.5%"),
        ("▶ 전국 급식 전체 합계 (최종)",
         total_c,       f"약 {total_c/1e8:.0f}억원",
         prev_total_c,  f"약 {prev_total_c/1e8:.0f}억원" if prev_total_c else "N/A",
         "전국 학교급식 + 단체급식 합산"),
    ]
    for i, (label, cv, cs, pv, ps, note) in enumerate(calc_rows, 11):
        ws2.row_dimensions[i].height = 28
        is_final = "▶" in label
        rf = F_ORANGE if is_final else F_WHITE
        rft = FT_BOLD if is_final else FT_N

        c2 = ws2.cell(i, 2, label)
        c2.fill = rf; c2.font = rft; c2.alignment = AL; c2.border = BD
        c3 = ws2.cell(i, 3, cv)
        c3.fill = rf; c3.font = rft; c3.alignment = AR; c3.border = BD; c3.number_format = "#,##0"
        c4 = ws2.cell(i, 4, cs)
        c4.fill = F_ORANGE if is_final else F_GREEN; c4.font = FT_SUM; c4.alignment = AC; c4.border = BD
        c5 = ws2.cell(i, 5, pv if pv else None)
        c5.fill = F_GRAY; c5.font = FT_N; c5.alignment = AR; c5.border = BD; c5.number_format = "#,##0"
        c6 = ws2.cell(i, 6, ps)
        c6.fill = F_GRAY; c6.font = FT_N; c6.alignment = AC; c6.border = BD
        c7 = ws2.cell(i, 7, note)
        c7.fill = F_GRAY; c7.font = FT_S; c7.alignment = AL; c7.border = BD

    # 방법론 설명
    ws2.row_dimensions[17].height = 10
    ws2.row_dimensions[18].height = 22
    ws2.merge_cells("B18:G18")
    mh = ws2.cell(18, 2, "환산 방법론 및 주의사항")
    mh.fill = F_LBLUE; mh.font = FT_LABEL; mh.alignment = AL; mh.border = BD

    methodology = [
        "• FoodnBid(info.foodnbid.com)는 수도권(서울·경기·인천) 학교급식 조달 플랫폼으로, 전체 수도권 학교급식의 약 60%를 커버합니다.",
        "• 수도권 학생 비율(40.5%)은 교육통계서비스 2024년 기준 초·중·고 재학생 수를 활용합니다.",
        "• 단체급식(군·병원·기업) 비중은 학교급식의 약 22.5%로 추정합니다 (업계 통상 기준).",
        "• 방학월(1·7·12월)은 학교 운영 없어 실측값이 매우 낮습니다 — 정상 데이터입니다.",
        f"• 본 데이터는 {year}년 및 {prev_year}년 FoodnBid 실측 기준이며, 연도별 편차가 존재할 수 있습니다.",
    ]
    for i, text in enumerate(methodology, 19):
        ws2.row_dimensions[i].height = 20
        ws2.merge_cells(f"B{i}:G{i}")
        c = ws2.cell(i, 2, text)
        c.fill = F_GRAY; c.font = FT_S; c.alignment = AL; c.border = BD

    # 출처
    ws2.row_dimensions[25].height = 14
    ws2.merge_cells("B25:G25")
    src2 = ws2.cell(25, 2,
        f"※ 출처: FoodnBid info.foodnbid.com | 교육통계서비스 kess.kedi.re.kr | 수집일: {datetime.now().strftime('%Y-%m-%d')}")
    src2.fill = F_GRAY; src2.font = FT_SRC; src2.alignment = AL; src2.border = BD

    # ════════════════════════════════════════════════════════
    #  SHEET 3 : 경쟁사_월별현황
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("경쟁사_월별현황")
    ws3.sheet_view.showGridLines = False

    # 전체 경쟁사 목록 집계
    comp_totals = {}
    for r in results:
        for co in r.get("companies", []):
            name = co["company"]
            comp_totals[name] = comp_totals.get(name, 0) + co["amount"]
    sorted_comps = sorted(comp_totals.items(), key=lambda x: -x[1])

    max_comps = min(len(sorted_comps), 10)
    comp_names = [n for n, _ in sorted_comps[:max_comps]]

    # 열 너비
    ws3.column_dimensions["A"].width = 2
    ws3.column_dimensions["B"].width = 8
    for j in range(max_comps):
        ws3.column_dimensions[get_column_letter(j+3)].width = 16
    ws3.column_dimensions[get_column_letter(max_comps+3)].width = 16

    # 제목
    ws3.row_dimensions[1].height = 38
    ws3.merge_cells(f"B1:{get_column_letter(max_comps+3)}1")
    c = ws3.cell(1, 2, f"【{keyword}】  경쟁사 월별 낙찰현황  ({year}년)")
    c.fill = F_BLUE; c.font = FT_TITLE; c.alignment = AC; c.border = BD

    # 헤더
    ws3.row_dimensions[2].height = 32
    ws3.cell(2, 2, "월").fill = F_BLUE
    ws3.cell(2, 2).font = FT_HEAD; ws3.cell(2, 2).alignment = AC; ws3.cell(2, 2).border = BD

    for j, name in enumerate(comp_names, 3):
        annual_share = round(comp_totals[name] / annual * 100, 1) if annual else 0
        c = ws3.cell(2, j, f"{name}\n({annual_share}%)")
        c.fill = F_KPINK if j == 3 else F_BLUE
        c.font = FT_HEAD_PINK if j == 3 else FT_HEAD
        c.alignment = AC; c.border = BD

    # 합계 열 헤더
    total_col = max_comps + 3
    c = ws3.cell(2, total_col, "월 합계")
    c.fill = F_GREEN; c.font = FT_SUM; c.alignment = AC; c.border = BD

    # 월별 데이터
    for r in results:
        row = r["month"] + 2
        ws3.row_dimensions[row].height = 20
        is_vac = r["month"] in VACATION
        row_fill = F_VAC if is_vac else F_WHITE

        # 월 레이블
        c2 = ws3.cell(row, 2, MONTH_NAMES[r["month"]-1])
        c2.fill = row_fill; c2.font = FT_BOLD if is_vac else FT_N
        c2.alignment = AC; c2.border = BD

        # 이 달의 경쟁사별 금액 조회
        month_map = {co["company"]: co["amount"] for co in r.get("companies", [])}

        for j, name in enumerate(comp_names, 3):
            amt = month_map.get(name, 0)
            # 이 달 1위인지 확인
            is_top = amt > 0 and amt == max(month_map.values(), default=0)
            c = ws3.cell(row, j, amt if amt else None)
            c.fill = F_KPINK if is_top else row_fill
            c.font = FT_BOLD if is_top else FT_N
            c.alignment = AR; c.border = BD
            c.number_format = "#,##0"

        # 월 합계
        ct = ws3.cell(row, total_col, r["total"])
        ct.fill = F_GREEN if not is_vac else F_VAC
        ct.font = FT_SUM; ct.alignment = AR; ct.border = BD; ct.number_format = "#,##0"

    # 연간 합계 행
    r_total = 15
    ws3.row_dimensions[r_total].height = 26
    c2 = ws3.cell(r_total, 2, "연간 합계")
    c2.fill = F_GREEN; c2.font = FT_SUM; c2.alignment = AC; c2.border = BD

    for j, (name, tot) in enumerate(sorted_comps[:max_comps], 3):
        share = round(tot / annual * 100, 1) if annual else 0
        c = ws3.cell(r_total, j, tot)
        c.fill = F_GREEN; c.font = FT_SUM; c.alignment = AR
        c.border = BD; c.number_format = "#,##0"
        # 점유율 sub-text
        ws3.row_dimensions[r_total+1].height = 16
        cs = ws3.cell(r_total+1, j, f"점유율 {share}%")
        cs.fill = F_LBLUE; cs.font = FT_S; cs.alignment = AC; cs.border = BD

    ct2 = ws3.cell(r_total, total_col, annual)
    ct2.fill = F_GREEN; ct2.font = FT_SUM; ct2.alignment = AR
    ct2.border = BD; ct2.number_format = "#,##0"

    # 출처
    ws3.row_dimensions[17].height = 14
    last_col_letter = get_column_letter(total_col)
    ws3.merge_cells(f"B17:{last_col_letter}17")
    src3 = ws3.cell(17, 2,
        f"※ 분홍색 셀 = 해당 월 1위 낙찰사  |  출처: FoodnBid info.foodnbid.com  |  수집일: {datetime.now().strftime('%Y-%m-%d')}")
    src3.fill = F_GRAY; src3.font = FT_SRC; src3.alignment = AL; src3.border = BD

    # ── 저장 ──────────────────────────────────────────────
    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    fname    = os.path.join(OUTPUT_DIR, f"{keyword}_학교급식규모_{year}_{date_str}.xlsx")
    wb.save(fname)
    print(f"\n  [엑셀] 저장 완료 → {fname}")
    return fname


def add_product_sheet(wb, keyword: str, year: int, product_info: list, bluesis_prices: dict = None):
    """
    Sheet 4: 경쟁사_제품정보
    FoodnBid에서 수집한 TOP5 경쟁사의 대표 제품 비교표를 생성합니다.
    - 업체명·제품명·중량·원산지: FoodnBid 자동 수집
    - FoodnBid 방학제외 월평균: 1·7·12월 제외 9개월 평균
    - 블루시스 납품가·제품 이미지·육함량: 직접 입력 필요 (placeholder)
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def s(): return Side(style="thin", color="CCCCCC")
    BD2 = Border(left=s(), right=s(), top=s(), bottom=s())

    AC2 = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL2 = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AR2 = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    P_BLUE   = PatternFill("solid", fgColor="1F4E79")
    P_LBLUE  = PatternFill("solid", fgColor="D6E4F0")
    P_LGREEN = PatternFill("solid", fgColor="E8F5E9")
    P_ORANGE = PatternFill("solid", fgColor="FFF3E0")
    P_GOLD   = PatternFill("solid", fgColor="FFFDE7")
    P_GRAY   = PatternFill("solid", fgColor="F0F0F0")
    P_WHITE  = PatternFill()
    P_RED    = PatternFill("solid", fgColor="FFE0E0")

    FH  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    FT2 = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
    FL  = Font(name="맑은 고딕", bold=True, color="1F4E79", size=10)
    FB  = Font(name="맑은 고딕", bold=True, size=10)
    FN2 = Font(name="맑은 고딕", size=10)
    FS2 = Font(name="맑은 고딕", size=9)
    FR  = Font(name="맑은 고딕", bold=True, color="C00000", size=10)
    FSRC2 = Font(name="맑은 고딕", size=8, italic=True, color="888888")
    FIMG = Font(name="맑은 고딕", size=9, italic=True, color="BBBBBB")

    RANK_COLORS = ["FF6B35", "FF9500", "4CAF50", "2196F3", "9C27B0"]

    if "경쟁사_제품정보" in wb.sheetnames:
        del wb["경쟁사_제품정보"]
    ws = wb.create_sheet("경쟁사_제품정보")

    n = len(product_info)
    last_col = get_column_letter(n + 2)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 15
    for j in range(n):
        ws.column_dimensions[get_column_letter(j + 3)].width = 24

    for r, h in [(1,40),(2,26),(3,90),(4,26),(5,56),(6,32),(7,36),(8,70),(9,28),(10,44),(11,44),(12,20)]:
        ws.row_dimensions[r].height = h

    # 1행 제목
    ws.merge_cells(f"B1:{last_col}1")
    c = ws.cell(1, 2, f"【{keyword}】  경쟁사 제품 비교  ({year}년 FoodnBid 기준)")
    c.fill = P_BLUE; c.font = FT2; c.alignment = AC2; c.border = BD2

    # 2행 순위 헤더
    lh = ws.cell(2, 2, "항목")
    lh.fill = P_BLUE; lh.font = FH; lh.alignment = AC2; lh.border = BD2
    for j, p in enumerate(product_info, 3):
        c = ws.cell(2, j, f"FoodnBid {year} 제품 {p['rank']}위")
        c.fill = PatternFill("solid", fgColor=RANK_COLORS[j-3])
        c.font = FH; c.alignment = AC2; c.border = BD2

    # 3행 이미지 placeholder
    li = ws.cell(3, 2, "제품\n이미지")
    li.fill = P_GRAY; li.font = FL; li.alignment = AC2; li.border = BD2
    for j in range(n):
        c = ws.cell(3, j+3, "[ 제품 이미지\n직접 삽입 ]")
        c.fill = PatternFill("solid", fgColor="F9F9F9")
        c.font = FIMG; c.alignment = AC2; c.border = BD2

    # 데이터 행 정의: (행번호, 라벨, 값함수, 정렬, 배경)
    def make_rows(p):
        bluesis_val = (bluesis_prices or {}).get(p["company"], "직접 확인 필요")
        return [
            (4,  "업체명",             p["company"],        AC2, None,     FH),
            (5,  "제품명",             p["product"] or "확인 필요", AC2, P_WHITE, FB),
            (6,  "중량",               p["weight"],          AC2, P_GRAY,  FN2),
            (7,  "FoodnBid\n방학제외 월평균", f"{p['monthly_avg']:,}원", AR2, P_LGREEN, FB),
            (8,  "원재료 TOP3\n(육함량 직접 기재)", _ingredient_hint(p), AL2, P_GOLD, FS2),
            (9,  "원산지",             p["origin"],          AC2, P_WHITE, FN2),
            (10, "블루시스\n학교kg단가", bluesis_val,          AC2, P_ORANGE, FS2),
            (11, "특이사항",           "",                   AL2, P_LBLUE, FS2),
        ]

    for j, p in enumerate(product_info, 3):
        rows = make_rows(p)
        for row_num, label, val, align, fill, font in rows:
            # 라벨
            if j == 3:
                lc = ws.cell(row_num, 2, label)
                lc.fill = P_BLUE if row_num == 4 else P_GRAY
                lc.font = FH if row_num == 4 else FL
                lc.alignment = AC2; lc.border = BD2

            row_fill = fill
            row_font = font
            if row_num == 4:
                row_fill = PatternFill("solid", fgColor=RANK_COLORS[j-3])
            if row_num == 9 and p["origin"] == "수입산":
                row_fill = P_RED
                row_font = FR

            c2 = ws.cell(row_num, j, val)
            c2.fill = row_fill if row_fill else P_WHITE
            c2.font = row_font; c2.alignment = align; c2.border = BD2

    # 12행 출처
    ws.merge_cells(f"B12:{last_col}12")
    src = ws.cell(12, 2,
        f"※ 출처: FoodnBid info.foodnbid.com  |  방학제외 = 1·7·12월 제외  "
        f"|  블루시스 학교kg단가: market.bluesis.com 자동 수집  |  이미지·육함량: 직접 입력  |  수집일: {datetime.now().strftime('%Y-%m-%d')}")
    src.font = FSRC2; src.alignment = AL2


def _ingredient_hint(p: dict) -> str:
    """원재료 힌트 (육함량·세부 성분은 직접 확인 필요)"""
    origin_tag = p["origin"] if p["origin"] != "확인 필요" else "원산지 직접확인"
    return (
        f"① 주원료({origin_tag}) - 육함량: 직접 기재\n"
        f"② 소스류(간장·설탕·전분 등) - 직접 기재\n"
        f"③ 향신료·기타 첨가물 - 직접 기재"
    )


# ── JSON 로그 저장 ────────────────────────────────────────
def save_log(keyword: str, year: int, results: list, exclude: str = ""):
    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    log_data = {
        "keyword": keyword,
        "year": year,
        "collected_at": date_str,
        "results": results,
        "annual_total": sum(r["total"] for r in results),
        "methodology": {
            "foodnbid_coverage": 0.60,
            "suwon_student_ratio": 0.405,
            "dangtche_ratio": 0.225,
            "military_excluded": True,
            "exclude_keyword": exclude
        }
    }
    fname = os.path.join(LOG_DIR, f"{keyword}_학교급식_{year}_{date_str}.json")
    with open(fname, "w", encoding="utf-8") as f:
        json.dump(log_data, f, ensure_ascii=False, indent=2)
    print(f"  [로그] 저장 완료 → {fname}")
    return fname


def save_product_log(keyword: str, year: int, product_info: list):
    """경쟁사 제품정보를 JSON으로 저장 (대시보드에서 읽어 사용)"""
    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    fname = os.path.join(LOG_DIR, f"{keyword}_제품정보_{year}_{date_str}.json")
    with open(fname, "w", encoding="utf-8") as f:
        json.dump(product_info, f, ensure_ascii=False, indent=2)
    print(f"  [로그] 제품정보 저장 완료 → {fname}")
    return fname


# ── 분석 기준 연도 (매년 1월에 수동 업데이트) ──────────────
ANALYSIS_YEAR = 2025   # 항상 이 연도 vs 전년도(2024) 비교


# ── 메인 ──────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    if not args:
        print("사용법: py -X utf8 학교급식규모.py <키워드>")
        print("예시:   py -X utf8 학교급식규모.py 새우만두")
        sys.exit(1)

    keyword = args[0]
    # 연도 인수가 있어도 ANALYSIS_YEAR 고정 (2025 vs 2024)
    year    = ANALYSIS_YEAR
    exclude = args[2] if len(args) > 2 else ""

    print("=" * 65)
    print(f"  학교급식 시장규모 스킬  |  키워드: {keyword}  |  {year}년 vs {year-1}년")
    if exclude:
        print(f"  검색제외 키워드: {exclude}")
    print("=" * 65)
    print()
    prev_year = year - 1
    print(f"[STEP 1] FoodnBid 로그인 및 {year}년 1~12월 데이터 수집...")
    results = asyncio.run(collect_annual(keyword, year, exclude))

    print()
    print(f"[STEP 1-2] {prev_year}년 전년도 비교 데이터 수집...")
    prev_results = asyncio.run(collect_annual(keyword, prev_year, exclude))

    # 진행중인 다음 연도(예: 2026) 데이터도 있으면 함께 수집
    next_year     = year + 1
    next_results  = None
    if datetime.now().year >= next_year:
        print()
        print(f"[STEP 1-3] {next_year}년 진행중 데이터 수집...")
        next_results = asyncio.run(collect_annual(keyword, next_year, exclude))

    print()
    print("[STEP 2] 결과 요약...")
    print_summary(keyword, year, results)

    print("[STEP 3] 로그 저장...")
    save_log(keyword, year, results, exclude)
    save_log(keyword, prev_year, prev_results, exclude)   # 전년도 로그도 자동 저장
    if next_results:
        save_log(keyword, next_year, next_results, exclude)   # 진행중 연도 로그 저장

    print("[STEP 4] 엑셀 저장...")
    fname = save_excel(keyword, year, results, prev_results)

    print()
    print("[STEP 5] 경쟁사 제품정보 수집 및 Sheet 4 생성...")
    product_info = asyncio.run(collect_top_products(keyword, year, results, exclude))

    if product_info:
        save_product_log(keyword, year, product_info)   # 대시보드용 JSON 저장

    bluesis_prices = {}
    if product_info:
        print()
        print("[STEP 6] 블루시스 학교kg단가 자동 수집...")
        bluesis_prices = asyncio.run(collect_bluesis_prices(keyword, product_info))

    if fname and product_info:
        import openpyxl
        wb2 = openpyxl.load_workbook(fname)
        add_product_sheet(wb2, keyword, year, product_info, bluesis_prices)
        wb2.save(fname)
        print(f"  [Sheet 4] 경쟁사_제품정보 추가 완료 (블루시스 가격 {sum(1 for v in bluesis_prices.values() if '확인 필요' not in v)}/{len(product_info)}건 수집)")

    if fname:
        import subprocess
        subprocess.Popen(["cmd", "/c", "start", "", fname])

    print("\n완료!")


if __name__ == "__main__":
    main()
