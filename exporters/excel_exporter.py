"""
엑셀 내보내기 모듈
여러 시트로 구성된 포맷된 엑셀 파일 생성
"""

import io
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

logger = logging.getLogger(__name__)

# ============================================================
# 색상 팔레트
# ============================================================
COLOR = {
    "header_blue": "1F4E79",
    "header_green": "1E6B3C",
    "header_orange": "833C00",
    "header_purple": "4B1F6F",
    "row_alt": "EBF5FB",
    "row_highlight": "FFF2CC",
    "white": "FFFFFF",
    "light_gray": "F2F2F2",
    "border": "BFBFBF",
}


def _header_fill(color_hex: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color_hex)


def _header_font(color_hex: str = "FFFFFF") -> Font:
    return Font(bold=True, color=color_hex, name="맑은 고딕", size=10)


def _cell_font(bold: bool = False) -> Font:
    return Font(bold=bold, name="맑은 고딕", size=10)


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _auto_width(ws, min_width: int = 8, max_width: int = 40):
    """열 너비 자동 조정"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                # 한글은 2배 폭
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                if char_len > max_len:
                    max_len = char_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def _format_header_row(ws, row: int, header_color: str, text_color: str = "FFFFFF"):
    """헤더 행 스타일 적용"""
    fill = _header_fill(header_color)
    font = _header_font(text_color)
    border = _thin_border()
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = _center()


def _format_data_rows(ws, start_row: int, end_row: int, alt_color: str = None):
    """데이터 행 스타일 적용"""
    border = _thin_border()
    for row_idx in range(start_row, end_row + 1):
        is_alt = (row_idx - start_row) % 2 == 1
        fill = PatternFill(fill_type="solid", fgColor=alt_color) if (alt_color and is_alt) else None
        for cell in ws[row_idx]:
            cell.border = border
            cell.font = _cell_font()
            if fill:
                cell.fill = fill
            # 숫자 오른쪽 정렬
            if isinstance(cell.value, (int, float)):
                cell.alignment = _right()
                if cell.value >= 1000:
                    cell.number_format = "#,##0"
            else:
                cell.alignment = _left()


# ============================================================
# 메인 엑셀 생성 함수
# ============================================================

def create_excel_report(
    sales_df: pd.DataFrame,
    product_df: pd.DataFrame,
    sales_type_df: pd.DataFrame,
    top_products_df: pd.DataFrame,
    regional_df: pd.DataFrame,
    competitive_analysis: dict,
    keyword: str,
    start_date: str,
    end_date: str,
) -> bytes:
    """
    전체 분석 결과를 엑셀 파일로 생성

    Returns:
        bytes: 엑셀 파일 바이트
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 시트 생성
    _create_summary_sheet(wb, sales_type_df, competitive_analysis, keyword, start_date, end_date)
    _create_sales_detail_sheet(wb, sales_df)
    _create_product_detail_sheet(wb, product_df)
    _create_top_products_sheet(wb, top_products_df)
    _create_regional_sheet(wb, regional_df)
    _create_competitive_sheet(wb, competitive_analysis, product_df)

    # 바이트로 저장
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _create_summary_sheet(wb, sales_type_df, competitive_analysis, keyword, start_date, end_date):
    """1. 요약 시트"""
    ws = wb.create_sheet("📊 요약 대시보드")
    ws.freeze_panes = "A4"

    # 타이틀
    ws.merge_cells("A1:H1")
    ws["A1"] = f"식품 시장 분석 보고서 | 키워드: [{keyword}] | 기간: {start_date} ~ {end_date}"
    ws["A1"].font = Font(bold=True, size=14, name="맑은 고딕", color=COLOR["header_blue"])
    ws["A1"].alignment = _center()
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="DEEAF1")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 수도권(서울/경기/인천) 기준"
    ws["A2"].font = Font(italic=True, size=9, name="맑은 고딕", color="595959")
    ws["A2"].alignment = _center()

    row = 4

    # 매출 타입별 요약
    if not sales_type_df.empty:
        ws.cell(row, 1, "▶ 제품 타입별 수도권 매출 순위").font = Font(bold=True, size=11, name="맑은 고딕")
        row += 1

        headers = ["순위", "제품타입", "총매출금액(원)", "총수량", "평균단가(원)", "제품수", "매출비중(%)"]
        for c, h in enumerate(headers, 1):
            ws.cell(row, c, h)
        _format_header_row(ws, row, COLOR["header_blue"])
        row += 1

        for rank, (_, r) in enumerate(sales_type_df.iterrows(), 1):
            ws.cell(row, 1, rank)
            ws.cell(row, 2, r.get("제품타입", ""))
            ws.cell(row, 3, r.get("총매출금액", 0))
            ws.cell(row, 4, r.get("총수량", 0))
            ws.cell(row, 5, r.get("평균단가", 0))
            ws.cell(row, 6, r.get("제품수", 0))
            ws.cell(row, 7, r.get("매출비중(%)", 0))
            row += 1

        _format_data_rows(ws, row - len(sales_type_df), row - 1, COLOR["row_alt"])
        row += 2

    # 경쟁력 분석 요약
    ws.cell(row, 1, "▶ 경쟁력 분석 요약 (블루시스마켓 기준)").font = Font(bold=True, size=11, name="맑은 고딕")
    row += 1

    summary_items = [
        ("학교매입가 중앙값", f"{competitive_analysis.get('median_price', 0):,}원"),
        ("경쟁력 가격대", f"{competitive_analysis.get('recommended_price_range', (0,0))[0]:,}원 ~ "
                       f"{competitive_analysis.get('recommended_price_range', (0,0))[1]:,}원"),
        ("권장 함량", f"{int(competitive_analysis.get('recommended_weight_g', 0))}g"),
        ("주력 TIER", competitive_analysis.get("dominant_tier", "-")),
        ("주요 조리법", competitive_analysis.get("dominant_cooking_method", "-")),
        ("주요 브랜드", ", ".join(competitive_analysis.get("top_brands", [])[:3]) or "-"),
        ("주요 제조사", ", ".join(competitive_analysis.get("top_manufacturers", [])[:3]) or "-"),
    ]

    for label, value in summary_items:
        ws.cell(row, 1, label).font = Font(bold=True, name="맑은 고딕", size=10)
        ws.cell(row, 1).fill = PatternFill(fill_type="solid", fgColor="DEEAF1")
        ws.cell(row, 2, value).font = _cell_font()
        for c in [1, 2]:
            ws.cell(row, c).border = _thin_border()
            ws.cell(row, c).alignment = _left()
        row += 1

    row += 1
    # 인사이트 텍스트
    insight = competitive_analysis.get("insight_text", "")
    if insight:
        ws.cell(row, 1, "▶ 기획 인사이트").font = Font(bold=True, size=11, name="맑은 고딕")
        row += 1
        ws.merge_cells(f"A{row}:H{row+8}")
        ws.cell(row, 1, insight.replace("**", "").replace("📌", "").replace("📦", "").replace("🏷", "").replace("🍳", "").replace("🏢", "").replace("💰", "").replace("🥇", "").replace("💡", ""))
        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row, 1).font = _cell_font()
        ws.row_dimensions[row].height = 120

    _auto_width(ws)
    ws.column_dimensions["B"].width = 45


def _create_sales_detail_sheet(wb, sales_df: pd.DataFrame):
    """2. 푸드앤비드 매출 상세"""
    ws = wb.create_sheet("📈 매출 상세 (푸드앤비드)")
    if sales_df.empty:
        ws["A1"] = "데이터 없음"
        return

    ws.freeze_panes = "A2"
    headers = list(sales_df.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    _format_header_row(ws, 1, COLOR["header_green"])

    for r_idx, row in enumerate(sales_df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(r_idx, c_idx, val)

    _format_data_rows(ws, 2, sales_df.shape[0] + 1, COLOR["row_alt"])
    _auto_width(ws)


def _create_product_detail_sheet(wb, product_df: pd.DataFrame):
    """3. 블루시스마켓 상품 상세"""
    ws = wb.create_sheet("🛒 상품 상세 (블루시스마켓)")
    if product_df.empty:
        ws["A1"] = "데이터 없음"
        return

    ws.freeze_panes = "A2"
    display_cols = ["키워드", "제품명", "학교매입가", "학교매입가_숫자", "함량", "함량_g",
                    "TIER", "조리법", "제조사", "브랜드", "상품URL"]
    cols = [c for c in display_cols if c in product_df.columns]
    df = product_df[cols]

    for c, h in enumerate(cols, 1):
        ws.cell(1, c, h)
    _format_header_row(ws, 1, COLOR["header_orange"])

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(r_idx, c_idx, val)

    _format_data_rows(ws, 2, df.shape[0] + 1, COLOR["row_alt"])
    _auto_width(ws)
    # URL 열은 좁게
    url_col_idx = cols.index("상품URL") + 1 if "상품URL" in cols else None
    if url_col_idx:
        ws.column_dimensions[get_column_letter(url_col_idx)].width = 30


def _create_top_products_sheet(wb, top_products_df: pd.DataFrame):
    """4. 매출 상위 제품"""
    ws = wb.create_sheet("🏆 매출 TOP20")
    if top_products_df.empty:
        ws["A1"] = "데이터 없음"
        return

    ws.freeze_panes = "A2"
    headers = ["순위"] + list(top_products_df.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    _format_header_row(ws, 1, COLOR["header_purple"])

    for r_idx, row in enumerate(top_products_df.itertuples(index=False), 2):
        ws.cell(r_idx, 1, r_idx - 1)  # 순위
        for c_idx, val in enumerate(row, 2):
            ws.cell(r_idx, c_idx, val)

    _format_data_rows(ws, 2, top_products_df.shape[0] + 1, COLOR["row_alt"])
    # Top 3 강조
    for rank_row in range(2, min(5, top_products_df.shape[0] + 2)):
        for cell in ws[rank_row]:
            cell.fill = PatternFill(fill_type="solid", fgColor=COLOR["row_highlight"])
    _auto_width(ws)


def _create_regional_sheet(wb, regional_df: pd.DataFrame):
    """5. 지역별 매출"""
    ws = wb.create_sheet("🗺 지역별 매출")
    if regional_df.empty:
        ws["A1"] = "데이터 없음"
        return

    ws.freeze_panes = "A2"
    headers = list(regional_df.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    _format_header_row(ws, 1, COLOR["header_green"])

    for r_idx, row in enumerate(regional_df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(r_idx, c_idx, val)

    _format_data_rows(ws, 2, regional_df.shape[0] + 1, COLOR["row_alt"])
    _auto_width(ws)


def _create_competitive_sheet(wb, competitive_analysis: dict, product_df: pd.DataFrame):
    """6. 경쟁력 분석 상세"""
    ws = wb.create_sheet("💡 경쟁력 분석")
    row = 1

    # TIER 분포
    tier_dist = competitive_analysis.get("tier_distribution", {})
    if tier_dist:
        ws.cell(row, 1, "TIER 분포").font = Font(bold=True, name="맑은 고딕")
        _format_header_row(ws, row, COLOR["header_blue"])
        row += 1
        for tier, cnt in tier_dist.items():
            ws.cell(row, 1, tier)
            ws.cell(row, 2, cnt)
            for c in [1, 2]:
                ws.cell(row, c).border = _thin_border()
                ws.cell(row, c).alignment = _left()
            row += 1
        row += 1

    # 조리법 분포
    cook_dist = competitive_analysis.get("cooking_method_distribution", {})
    if cook_dist:
        ws.cell(row, 1, "조리법 분포").font = Font(bold=True, name="맑은 고딕")
        _format_header_row(ws, row, COLOR["header_orange"])
        row += 1
        for method, cnt in cook_dist.items():
            ws.cell(row, 1, method)
            ws.cell(row, 2, cnt)
            for c in [1, 2]:
                ws.cell(row, c).border = _thin_border()
                ws.cell(row, c).alignment = _left()
            row += 1
        row += 1

    # 가격 분포 (함량별)
    if not product_df.empty and "학교매입가_숫자" in product_df.columns:
        ws.cell(row, 1, "학교매입가 분포").font = Font(bold=True, name="맑은 고딕")
        _format_header_row(ws, row, COLOR["header_purple"])
        row += 1
        price_bins = [0, 3000, 5000, 8000, 12000, 20000, 999999]
        price_labels = ["~3천원", "3~5천원", "5~8천원", "8~12천원", "12~20천원", "20천원~"]
        price_data = product_df["학교매입가_숫자"].replace(0, pd.NA).dropna()
        if not price_data.empty:
            counts = pd.cut(price_data, bins=price_bins, labels=price_labels).value_counts()
            for label, cnt in counts.items():
                ws.cell(row, 1, str(label))
                ws.cell(row, 2, int(cnt))
                for c in [1, 2]:
                    ws.cell(row, c).border = _thin_border()
                    ws.cell(row, c).alignment = _left()
                row += 1
        row += 1

    # 경쟁 상품 상세
    comp_products = competitive_analysis.get("competitive_products", pd.DataFrame())
    if not comp_products.empty:
        ws.cell(row, 1, "매출 TOP & 블루시스마켓 교차 분석").font = Font(bold=True, name="맑은 고딕")
        row += 1
        headers = list(comp_products.columns)
        for c, h in enumerate(headers, 1):
            ws.cell(row, c, h)
        _format_header_row(ws, row, COLOR["header_green"])
        row += 1
        for _, r in comp_products.iterrows():
            for c_idx, val in enumerate(r, 1):
                ws.cell(row, c_idx, val)
            row += 1
        _format_data_rows(ws, row - len(comp_products), row - 1)

    _auto_width(ws)
